# -*- coding: utf-8 -*-
"""
Versione Windows GUI convertita da notebook Colab Finale_codice_pl_1104.py.

Script Python standard per simulazione PL / imprevisti / recovery.

Esecuzione da PowerShell:
    python pl_pronto_intervento.py --input PL_1004.xlsx

Dipendenze:
    python -m pip install pandas numpy openpyxl
"""

import argparse
import math
import re
from copy import deepcopy
from collections import defaultdict
from functools import lru_cache
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ============================================================
# CONFIGURAZIONE UTENTE
# ============================================================

# Percorsi predefiniti. Puoi modificarli qui oppure passare --input e --output da terminale.
INPUT_XLSX = "PL_1004.xlsx"
OUTPUT_XLSX = "PL_1004_finale_pronto_intervento_output.xlsx"

# Numero di giorni lavorativi extra da aggiungere al calendario, per assorbire eventuali ritardi.
EXTRA_BUSINESS_DAYS = 60

# Attiva/disattiva scenari di recovery.
ENABLE_RECOVERY = True

# Se vuoi evitare le domande da terminale, inserisci qui gli imprevisti.
# Esempio:
# IMPREVISTI_INPUT = [
#     {
#         "Evento": "Pioggia",
#         "Data inizio": "2026-05-18",
#         "Ora inizio": "08:00-09:00",
#         "Data fine": "2026-05-18",
#         "Ora fine": "12:00-13:00",
#     }
# ]
IMPREVISTI_INPUT = []

RISK_LIBRARY = {
    "Rain": {
        "ID Rischio": "R1",
        "Evento": "Rain",
        "Tipo impatto": "Blocco totale",
        "Risorse colpite": "Tutte",
        "Recupero possibile?": "SI",
    },
    "High Wind": {
        "ID Rischio": "R2",
        "Evento": "High Wind",
        "Tipo impatto": "Blocco totale",
        "Risorse colpite": "Gru",
        "Recupero possibile?": "SI",
    },
    "Concrete Pump Failure": {
        "ID Rischio": "R3",
        "Evento": "Concrete Pump Failure",
        "Tipo impatto": "Blocco totale",
        "Risorse colpite": "Pompa Calcestruzzo",
        "Recupero possibile?": "SI",
    },
}

# Etichette che indicano tutte le risorse.
ALL_RESOURCES_LABELS = {"Tutte", "Tutti", "ALL", "All", "all", "tutte", "tutti"}

# ------------------------------------------------------------------
# GUI / OUTPUT TRANSLATIONS
# ------------------------------------------------------------------
# The simulation engine keeps the original Italian internal column names
# because the Excel model and calculations depend on them. These mappings
# are applied only to what the user sees in the GUI and to optional output
# workbook translation.
DISPLAY_VALUE_TRANSLATIONS_EN = {
    "Tutte": "All",
    "Tutti": "All",
    "tutte": "All",
    "tutti": "All",
    "SI": "YES",
    "NO": "NO",
    "Blocco totale": "Total shutdown",
    "Riduzione risorse": "Resource reduction",
    "Diretta": "Direct",
    "Indiretta": "Indirect",
    "Con imprevisti - senza recovery": "With risks - no recovery",
    "Con imprevisti + recovery diretta": "With risks + direct recovery",
    "Con imprevisti + recovery economica catena": "With risks + cost-optimized chain recovery",
    "Recovery diretta": "Direct recovery",
    "Recovery economica catena": "Cost-optimized chain recovery",
    "Gru": "Crane",
    "Pompa Calcestruzzo": "Concrete Pump",
    "Escavatore": "Excavator",
    "Autocarro": "Truck",
    "Operai Liv.1": "Workers Lv.1",
    "Operai Liv.2": "Workers Lv.2",
    "Operai Liv.3": "Workers Lv.3",
    "sempre": "always",
    "after_first_extra_team": "after first extra team",
}

OUTPUT_COLUMN_TRANSLATIONS_EN = {
    "ID Rischio": "Risk ID",
    "Evento": "Event",
    "Tipo impatto": "Impact type",
    "Risorse colpite": "Affected resources",
    "Recupero possibile?": "Recoverable?",
    "Data inizio": "Start date",
    "Ora inizio": "Start time",
    "Data fine": "End date",
    "Ora fine": "End time",
    "Risorsa": "Resource",
    "Start abs": "Start abs",
    "End abs": "End abs",
    "Ore impattate": "Impacted hours",
    "Cap media prima": "Average capacity before",
    "Cap media dopo": "Average capacity after",
    "Task colpita?": "Impacted task?",
    "Ore potenzialmente critiche": "Potentially critical hours",
    "Risorse critiche": "Critical resources",
    "Nome attività": "Activity name",
    "Elemento": "Element",
    "Ora assoluta Inizio": "Baseline start absolute hour",
    "Ora assoluta Fine": "Baseline finish absolute hour",
    "Nuova Ora assoluta Inizio": "New start absolute hour",
    "Nuova Ora assoluta Fine": "New finish absolute hour",
    "Ritardo fine (h)": "Finish delay (h)",
    "Ritardo fine (gg lav)": "Finish delay (work days)",
    "Predecessori": "Predecessors",
    "Scenario": "Scenario",
    "Fine progetto (ora assoluta)": "Project finish (h)",
    "Ritardo vs baseline (h)": "Delay vs baseline (h)",
    "Ritardo vs baseline (gg lav)": "Delay vs baseline (work days)",
    "Recovery attivo": "Recovery enabled",
    "Recovery eseguita": "Recovery executed",
    "Recovery logica": "Recovery logic",
    "Max extra team per task": "Max extra teams per task",
    "Min duration ratio": "Min duration ratio",
    "Costo totale baseline progetto": "Baseline total project cost",
    "Costo attività coinvolte": "Impacted activities cost",
    "Costo extra totale recovery": "Total recovery extra cost",
    "Costo totale progetto con recovery": "Project cost with recovery",
    "% extra su attività coinvolte": "% extra on impacted activities",
    "% extra su totale progetto": "% extra on total project",
    "Operai Liv.1": "Workers Lv.1",
    "Operai Liv.2": "Workers Lv.2",
    "Operai Liv.3": "Workers Lv.3",
    "Gru": "Crane",
    "Pompa Calcestruzzo": "Concrete Pump",
    "Escavatore": "Excavator",
    "Autocarro": "Truck",
    "Extra Operai Liv.1": "Extra Workers Lv.1",
    "Extra Operai Liv.2": "Extra Workers Lv.2",
    "Extra Operai Liv.3": "Extra Workers Lv.3",
    "Extra Gru": "Extra Crane",
    "Extra Pompa Calcestruzzo": "Extra Concrete Pump",
    "Extra Escavatore": "Extra Excavator",
    "Extra Autocarro": "Extra Truck",
    "Tipo impatto recovery": "Recovery impact type",
    "Regola macchinari": "Machinery rule",
    "Baseline multiplier": "Baseline multiplier",
    "Recovery multiplier": "Recovery multiplier",
    "Extra teams": "Extra teams",
    "Durata baseline": "Baseline duration (h)",
    "Durata recovery": "Recovery duration (h)",
    "Ore recovery": "Recovery hours",
    "Ore recuperate": "Recovered hours (h)",
    "Costo attività": "Activity cost",
    "Costo extra totale": "Total extra cost",
    "% costo extra totale": "% total extra cost",
    "Costo per ora recuperata": "Cost per recovered hour",
}

OUTPUT_SHEET_TRANSLATIONS_EN = {
    "PL_originale": "Original_PL",
    "Orario_originale": "Original_timetable",
    "Imprevisti_letti": "Risk_events",
    "Capacita_imprevisti": "Risk_capacities",
    "Log_imprevisti": "Risk_log",
    "Task_colpite": "Impacted_tasks",
    "Attivita_colpite": "Impacted_activities",
    "Task_bloccate": "Blocked_tasks",
    "Dettaglio_allocazione": "Allocation_detail",
    "PL_riprogrammato": "Rescheduled_PL",
    "Orario_riprogrammato": "Rescheduled_timetable",
    "Report_ritardi": "Delay_report",
    "Confronto_scenari": "Scenario_comparison",
    "Costo_recovery": "Recovery_cost",
    "Proposta_recovery_dir": "Direct_recovery_plan",
    "Capacita_recovery_dir": "Direct_recovery_capacity",
    "Recovery_log_dir": "Direct_recovery_log",
    "Dettaglio_recovery_dir": "Direct_recovery_detail",
    "Task_bloccate_dir": "Direct_blocked_tasks",
    "PL_recovery_dir": "Direct_recovery_PL",
    "Orario_recovery_dir": "Direct_recovery_timetable",
    "Report_recovery_dir": "Direct_recovery_report",
    "Candidate_chain": "Chain_candidates",
    "Proposta_chain": "Chain_plan",
    "Dettaglio_chain": "Chain_detail",
    "Task_bloccate_chain": "Chain_blocked_tasks",
    "PL_recovery_chain": "Chain_recovery_PL",
    "Orario_chain": "Chain_timetable",
    "Report_chain": "Chain_report",
}


def gui_display_value(value, language="EN"):
    if language != "EN":
        return value
    if value is None:
        return ""
    return DISPLAY_VALUE_TRANSLATIONS_EN.get(str(value), value)


def translate_visible_text(value, language="EN"):
    """Translate exact values and comma-separated/resource substrings for GUI display."""
    if language != "EN" or value is None:
        return value
    if not isinstance(value, str):
        return value
    out = DISPLAY_VALUE_TRANSLATIONS_EN.get(value, value)
    # Replace resource names inside lists such as "Gru, Operai Liv.1".
    for src, dst in sorted(DISPLAY_VALUE_TRANSLATIONS_EN.items(), key=lambda kv: len(kv[0]), reverse=True):
        out = out.replace(src, dst)
    return out


def postprocess_output_workbook_language(filepath, language="EN"):
    """Translate workbook labels/values after export. Internal calculations are already done."""
    if language != "EN":
        return filepath
    wb = load_workbook(filepath)

    for ws in wb.worksheets:
        # Translate sheet names safely after processing cells.
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    v = cell.value
                    if v in OUTPUT_COLUMN_TRANSLATIONS_EN:
                        cell.value = OUTPUT_COLUMN_TRANSLATIONS_EN[v]
                    else:
                        cell.value = translate_visible_text(v, "EN")

        new_title = OUTPUT_SHEET_TRANSLATIONS_EN.get(ws.title)
        if new_title:
            ws.title = new_title[:31]

    wb.save(filepath)
    return filepath



def translate_runtime_log_to_english(text):
    """Translate runtime console/log output generated by the Italian internal engine."""
    replacements = {
        # Section titles
        "===== AVVIO SIMULAZIONE =====": "===== SIMULATION START =====",
        "===== SIMULAZIONE COMPLETATA =====": "===== SIMULATION COMPLETED =====",
        "===== ERRORE DURANTE LA SIMULAZIONE =====": "===== ERROR DURING SIMULATION =====",
        "===== Anteprima PL =====": "===== PL PREVIEW =====",
        "===== IMPREVISTI INPUT =====": "===== RISK EVENTS INPUT =====",
        "===== LOG IMPREVISTI APPLICATI =====": "===== APPLIED RISK EVENTS LOG =====",
        "===== ATTIVITÀ DIRETTAMENTE COLPITE DAGLI IMPREVISTI =====": "===== ACTIVITIES DIRECTLY IMPACTED BY RISK EVENTS =====",
        "===== CONFRONTO SCENARI =====": "===== SCENARIO COMPARISON =====",
        "===== COSTI RECOVERY - RIEPILOGO =====": "===== RECOVERY COST SUMMARY =====",
        "===== DETTAGLIO COSTI - RECOVERY DIRETTA =====": "===== DIRECT RECOVERY COST DETAILS =====",
        "===== EXTRA RISORSE - RECOVERY DIRETTA =====": "===== EXTRA RESOURCES - DIRECT RECOVERY =====",
        "===== DETTAGLIO COSTI - RECOVERY ECONOMICA CATENA =====": "===== COST-OPTIMIZED CHAIN RECOVERY COST DETAILS =====",
        "===== EXTRA RISORSE - RECOVERY ECONOMICA CATENA =====": "===== EXTRA RESOURCES - COST-OPTIMIZED CHAIN RECOVERY =====",
        "===== CANDIDATE RECOVERY ECONOMICA SU CATENA =====": "===== COST-OPTIMIZED CHAIN RECOVERY CANDIDATES =====",
        # Sentences / labels
        "Input caricati.": "Input loaded.",
        "Orario:": "Timetable:",
        "Colonna ID attività PL:": "PL task ID column:",
        "Colonna ID attività Orario:": "Timetable task ID column:",
        "Risorse operative:": "Operational resources:",
        "Imprevisti:": "Risk events:",
        "File creato:": "File created:",
        "Nessun costo recovery da stampare.": "No recovery cost to display.",
        "Recovery non eseguita perché non c'è ritardo finale.": "Recovery was not executed because there is no final delay.",
        "Non riesco a leggere il confronto scenari dall'output:": "Could not read the scenario comparison from the output:",
        # Column / dataframe headers and values
        "ID Rischio": "Risk ID",
        "Evento": "Event",
        "Tipo impatto": "Impact type",
        "Risorse colpite": "Affected resources",
        "Recupero possibile?": "Recoverable?",
        "Data inizio": "Start date",
        "Ora inizio": "Start time",
        "Data fine": "End date",
        "Ora fine": "End time",
        "Risorsa": "Resource",
        "Ore impattate": "Impacted hours",
        "Cap media prima": "Average capacity before",
        "Cap media dopo": "Average capacity after",
        "Task colpita?": "Impacted task?",
        "Ore potenzialmente critiche": "Potentially critical hours",
        "Risorse critiche": "Critical resources",
        "Criticality": "Criticality",
        "Elemento": "Element",
        "Nome attività": "Activity name",
        "Durata\\n(h arrot)": "Duration\\n(rounded h)",
        "Durata\\n(h)": "Duration\\n(h)",
        "Durata": "Duration",
        "Ora assoluta Inizio": "Baseline start absolute hour",
        "Ora assoluta Fine": "Baseline finish absolute hour",
        "Predecessori": "Predecessors",
        "Predecessori (effettivi)": "Effective predecessors",
        "Nomi risorse": "Resource names",
        "Moltiplicatore risorse": "Resource multiplier",
        "Costo": "Cost",
        "Note": "Notes",
        "full_name": "full_name",
        "duration_hours": "duration_hours",
        "task_id": "task_id",
        "Scenario": "Scenario",
        "Fine progetto (ora assoluta)": "Project finish (h)",
        "Ritardo vs baseline (h)": "Delay vs baseline (h)",
        "Ritardo vs baseline (gg lav)": "Delay vs baseline (work days)",
        "Costo totale baseline progetto": "Baseline total project cost",
        "Costo attività coinvolte": "Impacted activities cost",
        "Costo extra totale recovery": "Total recovery extra cost",
        "Costo totale progetto con recovery": "Project cost with recovery",
        "% extra su attività coinvolte": "% extra on impacted activities",
        "% extra su totale progetto": "% extra on total project",
        # Values
        "Blocco totale": "Total shutdown",
        "Riduzione risorse": "Resource reduction",
        "Tutte": "All",
        "Tutti": "All",
        "Con imprevisti - senza recovery": "With risks - no recovery",
        "Con imprevisti + recovery diretta": "With risks + direct recovery",
        "Con imprevisti + recovery economica catena": "With risks + cost-optimized chain recovery",
        "Recovery diretta": "Direct recovery",
        "Recovery economica catena": "Cost-optimized chain recovery",
        "Diretta": "Direct",
        "Indiretta": "Indirect",
        "SI": "YES",
        # Resource labels for log readability
        "Operai Liv.1": "Workers Lv.1",
        "Operai Liv.2": "Workers Lv.2",
        "Operai Liv.3": "Workers Lv.3",
        "Gru": "Crane",
        "Pompa Calcestruzzo": "Concrete Pump",
        "Escavatore": "Excavator",
        "Autocarro": "Truck",
    }
    # Longer keys first prevents partial replacements from breaking longer labels.
    for old in sorted(replacements, key=len, reverse=True):
        text = text.replace(old, replacements[old])
    return text

# ============================================================
# CONFIG RECOVERY A SQUADRE
# ============================================================

ORE_GIORNO = 8

COSTI_OPERAI_ORARI = {
    "Operai Liv.1": 26,
    "Operai Liv.2": 29,
    "Operai Liv.3": 31,
}

COSTI_MACCHINARI = {
    "Gru": {"tipo_costo": "giornaliero", "costo": 300},
    "Pompa Calcestruzzo": {"tipo_costo": "orario", "costo": 210},
    "Escavatore": {"tipo_costo": "orario", "costo": 0},
    "Autocarro": {"tipo_costo": "orario", "costo": 0},
}

MAX_EXTRA_TEAMS_PER_TASK = 1

RECOVERABLE_ACTIVITIES = {
    "Armatura",
    "Getto",
    "Posa",
    "Casseratura Orizzontale",
    "Casseratura Verticale",
    "Scasseratura Orizzontale",
    "Scasseratura Verticale",
}

NON_RECOVERABLE_ACTIVITIES = {"Presa"}
MIN_TASK_DURATION_FOR_COMPRESSION = 2
MIN_DURATION_RATIO = 0.50

MACHINE_SCALE_RULES = {
    "Getto": "always",
    "Posa": "always",
    "Armatura": "after_first_extra_team",
    "Casseratura Orizzontale": "after_first_extra_team",
    "Casseratura Verticale": "after_first_extra_team",
    "Scasseratura Orizzontale": "after_first_extra_team",
    "Scasseratura Verticale": "after_first_extra_team",
    "Presa": "never",
}

# Queste variabili vengono valorizzate automaticamente leggendo il file Excel.
LABOR_COLS = []
MACHINE_COLS = []
RESOURCE_COLS = []


def show_df(df, name=None, max_rows=20):
    """Sostituisce display() di Colab/Jupyter con una stampa compatibile con terminale."""
    if name:
        print(f"\n===== {name} =====")
    if df is None:
        print("<None>")
        return
    if getattr(df, "empty", False):
        print("<DataFrame vuoto>")
        return
    with pd.option_context("display.max_rows", max_rows, "display.max_columns", None, "display.width", 180):
        print(df.head(max_rows).to_string(index=False))
        if len(df) > max_rows:
            print(f"... ({len(df) - max_rows} righe non mostrate)")


def display(obj):
    """Compatibilità minima con il codice esportato da notebook."""
    if isinstance(obj, pd.DataFrame):
        show_df(obj)
    else:
        print(obj)


def euro_it(x):
    return f"€ {float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def pct_it(x):
    return f"{float(x):.2%}"


def infer_resource_columns(pl_df):
    """Individua le colonne risorse attese nel foglio PL."""
    labor = [c for c in COSTI_OPERAI_ORARI if c in pl_df.columns]
    machines = [c for c in COSTI_MACCHINARI if c in pl_df.columns]
    missing_labor = [c for c in COSTI_OPERAI_ORARI if c not in pl_df.columns]
    missing_machines = [c for c in COSTI_MACCHINARI if c not in pl_df.columns]

    if missing_labor:
        print("Attenzione: colonne operai non trovate nel PL:", missing_labor)
    if missing_machines:
        print("Attenzione: colonne macchinari non trovate nel PL:", missing_machines)
    if not labor and not machines:
        raise ValueError("Non ho trovato colonne risorsa nel foglio PL. Controlla i nomi delle colonne.")
    return labor, machines, labor + machines


def find_column(columns, contains, required=True):
    matches = [c for c in columns if contains.lower() in str(c).lower()]
    if not matches and required:
        raise ValueError(f"Colonna contenente '{contains}' non trovata. Colonne disponibili: {list(columns)}")
    return matches[0] if matches else None


def normalize_input_risk(row):
    """Completa una riga imprevisto usando RISK_LIBRARY quando Evento corrisponde a un rischio noto."""
    row = dict(row)
    evento = str(row.get("Evento", row.get("Rischio", ""))).strip()
    if evento in RISK_LIBRARY:
        merged = RISK_LIBRARY[evento].copy()
        merged.update(row)
        return merged
    required = ["ID Rischio", "Evento", "Tipo impatto", "Risorse colpite", "Recupero possibile?"]
    missing = [c for c in required if c not in row or str(row[c]).strip() == ""]
    if missing:
        raise ValueError(f"Imprevisto personalizzato incompleto. Mancano: {missing}")
    return row


def build_imprevisti_df(input_rows=None):
    rows = IMPREVISTI_INPUT if input_rows is None else input_rows
    rows = [normalize_input_risk(r) for r in rows]
    df = pd.DataFrame(rows)
    if not df.empty:
        df.columns = [str(c).strip() for c in df.columns]
    return df


def prompt_imprevisti(available_hours):
    """Inserimento imprevisti da terminale classico, senza interfaccia grafica."""
    rows = []
    risk_names = list(RISK_LIBRARY.keys())
    print("\nInserimento imprevisti da terminale.")
    print("Rischi disponibili:")
    for i, name in enumerate(risk_names, start=1):
        print(f"  {i}) {name}")
    print("Ore disponibili nel calendario:", ", ".join(map(str, available_hours[:12])), "..." if len(available_hours) > 12 else "")

    while True:
        choice = input("\nNumero rischio, nome rischio, oppure INVIO per terminare: ").strip()
        if not choice:
            break
        if choice.isdigit() and 1 <= int(choice) <= len(risk_names):
            risk_name = risk_names[int(choice) - 1]
        else:
            risk_name = choice
        if risk_name not in RISK_LIBRARY:
            print(f"Rischio non valido: {risk_name}")
            continue

        row = RISK_LIBRARY[risk_name].copy()
        row["Data inizio"] = input("Data inizio (YYYY-MM-DD): ").strip()
        row["Ora inizio"] = input("Ora inizio, es. 08:00-09:00: ").strip()
        row["Data fine"] = input("Data fine (YYYY-MM-DD): ").strip()
        row["Ora fine"] = input("Ora fine, es. 12:00-13:00: ").strip()
        rows.append(row)
        print("Imprevisto aggiunto.")
    return rows


def prompt_imprevisti_gui(available_hours, available_dates):
    """
    Interfaccia grafica esterna con menu a tendina per inserire gli imprevisti.

    Usa tkinter, incluso nell'installazione standard di Python su Windows.
    Restituisce una lista di righe compatibile con build_imprevisti_df().
    """
    try:
        import tkinter as tk
        from tkinter import ttk, messagebox
    except Exception as exc:
        print("Impossibile avviare la finestra grafica tkinter:", exc)
        print("Passo all'inserimento classico da terminale.")
        return prompt_imprevisti(available_hours)

    rows = []
    risk_names = list(RISK_LIBRARY.keys())

    # Valori iniziali comodi.
    default_date = available_dates[0] if available_dates else ""
    default_hour_start = available_hours[0] if available_hours else ""
    default_hour_end = available_hours[-1] if available_hours else ""

    root = tk.Tk()
    root.title("Input Imprevisti")
    root.geometry("980x430")
    root.minsize(860, 390)

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    main = ttk.Frame(root, padding=14)
    main.pack(fill="both", expand=True)

    title = ttk.Label(main, text="Input Imprevisti", font=("Segoe UI", 14, "bold"))
    title.grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 12))

    ttk.Label(main, text="Rischio:").grid(row=1, column=0, sticky="e", padx=(0, 8), pady=4)
    risk_var = tk.StringVar(value=risk_names[0] if risk_names else "")
    risk_combo = ttk.Combobox(main, textvariable=risk_var, values=risk_names, state="readonly", width=32)
    risk_combo.grid(row=1, column=1, sticky="w", pady=4)

    ttk.Label(main, text="Data inizio:").grid(row=2, column=0, sticky="e", padx=(0, 8), pady=4)
    start_date_var = tk.StringVar(value=default_date)
    start_date_combo = ttk.Combobox(main, textvariable=start_date_var, values=available_dates, state="readonly", width=32)
    start_date_combo.grid(row=2, column=1, sticky="w", pady=4)

    ttk.Label(main, text="Ora inizio:").grid(row=2, column=2, sticky="e", padx=(24, 8), pady=4)
    start_hour_var = tk.StringVar(value=default_hour_start)
    start_hour_combo = ttk.Combobox(main, textvariable=start_hour_var, values=available_hours, state="readonly", width=24)
    start_hour_combo.grid(row=2, column=3, sticky="w", pady=4)

    ttk.Label(main, text="Data fine:").grid(row=3, column=0, sticky="e", padx=(0, 8), pady=4)
    end_date_var = tk.StringVar(value=default_date)
    end_date_combo = ttk.Combobox(main, textvariable=end_date_var, values=available_dates, state="readonly", width=32)
    end_date_combo.grid(row=3, column=1, sticky="w", pady=4)

    ttk.Label(main, text="Ora fine:").grid(row=3, column=2, sticky="e", padx=(24, 8), pady=4)
    end_hour_var = tk.StringVar(value=default_hour_end)
    end_hour_combo = ttk.Combobox(main, textvariable=end_hour_var, values=available_hours, state="readonly", width=24)
    end_hour_combo.grid(row=3, column=3, sticky="w", pady=4)

    button_frame = ttk.Frame(main)
    button_frame.grid(row=4, column=0, columnspan=4, sticky="w", pady=(10, 10))

    columns = (
        "ID Rischio",
        "Evento",
        "Tipo impatto",
        "Risorse colpite",
        "Recupero possibile?",
        "Data inizio",
        "Data fine",
        "Ora inizio",
        "Ora fine",
    )

    table = ttk.Treeview(main, columns=columns, show="headings", height=7)
    for col in columns:
        table.heading(col, text=col)
        width = 115
        if col in {"Tipo impatto", "Risorse colpite", "Recupero possibile?"}:
            width = 145
        if col in {"Data inizio", "Data fine", "Ora inizio", "Ora fine"}:
            width = 105
        table.column(col, width=width, anchor="center")

    table.grid(row=5, column=0, columnspan=4, sticky="nsew", pady=(0, 8))
    scrollbar = ttk.Scrollbar(main, orient="vertical", command=table.yview)
    scrollbar.grid(row=5, column=4, sticky="ns")
    table.configure(yscrollcommand=scrollbar.set)

    status_var = tk.StringVar(value="Aggiungi uno o più imprevisti, poi clicca Esegui simulazione.")
    status = ttk.Label(main, textvariable=status_var)
    status.grid(row=6, column=0, columnspan=4, sticky="w")

    main.columnconfigure(1, weight=1)
    main.columnconfigure(3, weight=1)
    main.rowconfigure(5, weight=1)

    def refresh_table():
        for item in table.get_children():
            table.delete(item)
        for row in rows:
            table.insert("", "end", values=[row.get(c, "") for c in columns])

    def add_row():
        risk_name = risk_var.get().strip()
        if risk_name not in RISK_LIBRARY:
            messagebox.showerror("Error", "Select a valid risk event.")
            return

        data_inizio = start_date_var.get().strip()
        data_fine = end_date_var.get().strip()
        ora_inizio = start_hour_var.get().strip()
        ora_fine = end_hour_var.get().strip()

        if not all([data_inizio, data_fine, ora_inizio, ora_fine]):
            messagebox.showerror("Errore", "Compila data e ora di inizio/fine.")
            return

        row = RISK_LIBRARY[risk_name].copy()
        row["Data inizio"] = data_inizio
        row["Data fine"] = data_fine
        row["Ora inizio"] = ora_inizio
        row["Ora fine"] = ora_fine
        rows.append(row)
        refresh_table()
        status_var.set(f"Risk events added: {len(rows)}")

    def remove_selected():
        selected = table.selection()
        if not selected:
            messagebox.showinfo("No selection", "Select a row to remove.")
            return
        indices = sorted((table.index(item) for item in selected), reverse=True)
        for idx in indices:
            if 0 <= idx < len(rows):
                rows.pop(idx)
        refresh_table()
        status_var.set(f"Risk events added: {len(rows)}")

    def run_and_close():
        if not rows:
            messagebox.showerror("Errore", "Add at least one risk event before running the simulation.")
            return
        root.destroy()

    def cancel():
        rows.clear()
        root.destroy()

    add_btn = ttk.Button(button_frame, text="Add risk event", command=add_row)
    add_btn.pack(side="left", padx=(0, 8))

    remove_btn = ttk.Button(button_frame, text="Remove selected", command=remove_selected)
    remove_btn.pack(side="left", padx=(0, 8))

    run_btn = ttk.Button(button_frame, text="Esegui simulazione", command=run_and_close)
    run_btn.pack(side="left", padx=(18, 8))

    cancel_btn = ttk.Button(button_frame, text="Annulla", command=cancel)
    cancel_btn.pack(side="left")

    root.bind("<Return>", lambda _event: add_row())
    root.protocol("WM_DELETE_WINDOW", cancel)

    root.mainloop()
    return rows

def parse_predecessors(x):
    if pd.isna(x):
        return []
    parts = re.split(r"[;,]+", str(x))
    return [int(p.strip()) for p in parts if p.strip().isdigit()]


def parse_id_list(x):
    if pd.isna(x):
        return []
    parts = re.split(r"[;,]+", str(x))
    return [int(p.strip()) for p in parts if p.strip().isdigit()]


def normalize_yes_no(v):
    return str(v).strip().upper() == "SI"


def clone_tasks(tasks):
    return deepcopy(tasks)


def normalize_hour_label(s):
    s = str(s).strip()
    m = re.match(r"^\s*(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})\s*$", s)
    if not m:
        return s
    h1, m1, h2, m2 = m.groups()
    return f"{int(h1):02d}:{m1}-{int(h2):02d}:{m2}"


def build_tasks(pl_df):
    pl = pl_df.copy()

    pl["task_id"] = pd.to_numeric(pl["task_id"], errors="coerce").astype(int)
    pl["duration_hours"] = pd.to_numeric(pl["duration_hours"], errors="coerce").fillna(0).astype(int)
    pl["Ora assoluta Inizio"] = pd.to_numeric(pl["Ora assoluta Inizio"], errors="coerce").astype(int)
    pl["Ora assoluta Fine"] = pd.to_numeric(pl["Ora assoluta Fine"], errors="coerce").astype(int)
    pl["Predecessori"] = pl["Predecessori"].apply(parse_predecessors)

    if "Moltiplicatore risorse" not in pl.columns:
        raise ValueError("Manca la colonna 'Moltiplicatore risorse' nel foglio PL.")

    pl["Moltiplicatore risorse"] = pd.to_numeric(
        pl["Moltiplicatore risorse"],
        errors="coerce"
    ).fillna(1).astype(float)

    for c in LABOR_COLS:
        pl[c] = pd.to_numeric(pl[c], errors="coerce").fillna(0.0)

    # salvo i flag macchina SI/NO
    for c in MACHINE_COLS:
        pl[f"{c}_flag"] = pl[c].apply(normalize_yes_no)

    # SI => quantità macchina = moltiplicatore risorse
    # NO => 0
    for c in MACHINE_COLS:
        pl[c] = pl.apply(
            lambda row: float(row["Moltiplicatore risorse"]) if row[f"{c}_flag"] else 0.0,
            axis=1
        )

    succ = defaultdict(list)
    dur_map = pl.set_index("task_id")["duration_hours"].to_dict()

    for _, row in pl.iterrows():
        for p in row["Predecessori"]:
            succ[p].append(int(row["task_id"]))

    @lru_cache(None)
    def critical_tail(task_id):
        if task_id not in succ or len(succ[task_id]) == 0:
            return dur_map[task_id]
        return dur_map[task_id] + max(critical_tail(s) for s in succ[task_id])

    tasks = {}

    for _, row in pl.iterrows():
        tid = int(row["task_id"])
        base_demand = {c: float(row[c]) for c in RESOURCE_COLS}

        machine_flags = {
            c: bool(row[f"{c}_flag"])
            for c in MACHINE_COLS
        }

        tasks[tid] = {
            "id": tid,
            "name": str(row["Nome attività"]).strip(),
            "full_name": str(row["full_name"]).strip(),
            "elemento": str(row["Elemento"]).strip(),

            "duration": int(row["duration_hours"]),
            "base_cost": float(row["Costo"]),

            "baseline_multiplier": float(row["Moltiplicatore risorse"]),
            "recovery_multiplier": float(row["Moltiplicatore risorse"]),

            "base_start": int(row["Ora assoluta Inizio"]),
            "base_finish": int(row["Ora assoluta Fine"]),
            "preds": list(row["Predecessori"]),
            "criticality": float(critical_tail(tid)),

            "machine_flags": machine_flags,

            "base_demand": base_demand,
            "effective_duration": int(row["duration_hours"]),
            "effective_demand": dict(base_demand),

            "remaining": int(row["duration_hours"]),
            "start": None,
            "finish": None,
            "assigned_hours": [],
        }

    return pl, tasks


def build_extended_calendar(orario_df, extra_business_days=10):
    cal = orario_df[["Data", "Ora", "Ora Assoluta"]].copy()
    cal["Data"] = pd.to_datetime(cal["Data"]).dt.normalize()
    cal["Ora"] = cal["Ora"].apply(normalize_hour_label)
    cal["Ora Assoluta"] = pd.to_numeric(cal["Ora Assoluta"], errors="coerce").astype(int)

    cal = cal.drop_duplicates("Ora Assoluta").sort_values("Ora Assoluta").reset_index(drop=True)

    first_day = cal["Data"].min()
    day_pattern = cal.loc[cal["Data"] == first_day, "Ora"].astype(str).tolist()

    last_abs = int(cal["Ora Assoluta"].max())
    last_date = pd.Timestamp(cal["Data"].max())

    extra_rows = []
    current_abs = last_abs
    current_date = last_date

    for _ in range(extra_business_days):
        current_date = pd.Timestamp(current_date) + pd.tseries.offsets.BDay(1)
        for hour_label in day_pattern:
            current_abs += 1
            extra_rows.append({
                "Data": pd.Timestamp(current_date).normalize(),
                "Ora": str(hour_label),
                "Ora Assoluta": current_abs
            })

    if extra_rows:
        cal = pd.concat([cal, pd.DataFrame(extra_rows)], ignore_index=True)

    cal = cal.sort_values("Ora Assoluta").reset_index(drop=True)
    return cal


def compute_original_peak_capacities(orario_df, pl_clean, id_col):
    task_res = pl_clean.set_index("task_id")[RESOURCE_COLS]
    rows = []

    for _, row in orario_df[["Ora Assoluta", id_col]].iterrows():
        ids = parse_id_list(row[id_col])

        if ids:
            ser = task_res.loc[ids].sum()
        else:
            ser = pd.Series({c: 0.0 for c in RESOURCE_COLS})

        ser["Ora Assoluta"] = int(row["Ora Assoluta"])
        rows.append(ser)

    usage_df = pd.DataFrame(rows).fillna(0.0)
    return usage_df[RESOURCE_COLS].max().to_dict()


def preprocess_imprevisti(imprevisti_df):
    imp = imprevisti_df.copy()
    imp["Data inizio"] = pd.to_datetime(imp["Data inizio"]).dt.normalize()
    imp["Data fine"] = pd.to_datetime(imp["Data fine"]).dt.normalize()
    imp["Ora inizio"] = imp["Ora inizio"].apply(normalize_hour_label)
    imp["Ora fine"] = imp["Ora fine"].apply(normalize_hour_label)
    imp["Tipo impatto"] = imp["Tipo impatto"].astype(str).str.strip()
    imp["Risorse colpite"] = imp["Risorse colpite"].astype(str).str.strip()
    imp["Recupero possibile?"] = imp["Recupero possibile?"].astype(str).str.strip().str.upper()

    if "Intensità" not in imp.columns:
        imp["Intensità"] = np.nan

    imp["Intensità"] = pd.to_numeric(imp["Intensità"], errors="coerce")
    imp["Intensità"] = np.where(
        imp["Tipo impatto"].str.upper().eq("BLOCCO TOTALE"),
        1.0,
        imp["Intensità"]
    )
    imp["Intensità"] = pd.to_numeric(imp["Intensità"], errors="coerce").fillna(1.0).clip(0, 1)

    return imp


def make_calendar_key_map(calendar_df):
    return {
        (row["Data"], str(row["Ora"]).strip()): int(row["Ora Assoluta"])
        for _, row in calendar_df.iterrows()
    }


def build_capacity_table_from_imprevisti(calendar_df, peak_caps, imprevisti_df):
    capacities = calendar_df.copy()
    capacities["Data"] = pd.to_datetime(capacities["Data"]).dt.normalize()
    capacities["Ora"] = capacities["Ora"].apply(normalize_hour_label)

    for r, cap in peak_caps.items():
        capacities[r] = float(cap)

    log_rows = []

    for _, imp in imprevisti_df.iterrows():

        start_date = pd.Timestamp(imp["Data inizio"]).normalize()
        end_date = pd.Timestamp(imp["Data fine"]).normalize()

        start_hour = normalize_hour_label(imp["Ora inizio"])
        end_hour = normalize_hour_label(imp["Ora fine"])

        tipo = str(imp["Tipo impatto"]).strip().upper()
        intensity = float(imp["Intensità"])

        raw_res = str(imp["Risorse colpite"]).strip()

        if raw_res in ALL_RESOURCES_LABELS:
            target_resources = list(RESOURCE_COLS)
        else:
            target_resources = [x.strip() for x in re.split(r"[;,]+", raw_res) if x.strip()]

        # giorni presenti nel calendario dentro il range imprevisto
        valid_days = sorted(
            capacities.loc[
                (capacities["Data"] >= start_date) &
                (capacities["Data"] <= end_date),
                "Data"
            ].drop_duplicates()
        )

        if len(valid_days) == 0:
            raise ValueError(
                f"Nessun giorno calendario trovato per imprevisto {imp['ID Rischio']} "
                f"tra {start_date.date()} e {end_date.date()}"
            )

        for day in valid_days:

            day_rows = capacities[capacities["Data"] == day].copy()

            if start_hour not in set(day_rows["Ora"]):
                raise ValueError(
                    f"Ora inizio {start_hour} non trovata nel calendario per il giorno {day.date()}"
                )

            if end_hour not in set(day_rows["Ora"]):
                raise ValueError(
                    f"Ora fine {end_hour} non trovata nel calendario per il giorno {day.date()}"
                )

            start_abs = int(day_rows.loc[day_rows["Ora"] == start_hour, "Ora Assoluta"].iloc[0])
            end_abs = int(day_rows.loc[day_rows["Ora"] == end_hour, "Ora Assoluta"].iloc[0])

            if end_abs < start_abs:
                raise ValueError(
                    f"Imprevisto {imp['ID Rischio']} ha end_abs < start_abs nel giorno {day.date()}"
                )

            for r in target_resources:
                if r not in capacities.columns:
                    print(f"Attenzione: risorsa '{r}' non trovata. Salto.")
                    continue

                mask = (
                    (capacities["Ora Assoluta"] >= start_abs) &
                    (capacities["Ora Assoluta"] <= end_abs)
                )

                before = capacities.loc[mask, r].copy()

                if tipo == "BLOCCO TOTALE":
                    capacities.loc[mask, r] = 0.0
                else:
                    capacities.loc[mask, r] = capacities.loc[mask, r] * (1.0 - intensity)

                after = capacities.loc[mask, r]

                log_rows.append({
                    "ID Rischio": imp["ID Rischio"],
                    "Evento": imp["Evento"],
                    "Data": day,
                    "Risorsa": r,
                    "Start abs": start_abs,
                    "End abs": end_abs,
                    "Ore impattate": end_abs - start_abs + 1,
                    "Cap media prima": float(before.mean()) if len(before) else np.nan,
                    "Cap media dopo": float(after.mean()) if len(after) else np.nan,
                    "Recupero possibile?": imp["Recupero possibile?"]
                })

    return capacities, pd.DataFrame(log_rows)

def identify_impacted_tasks(tasks, capacities_df):
    cap = capacities_df.set_index("Ora Assoluta")
    rows = []

    for tid, task in tasks.items():
        demanded = {r: v for r, v in task["base_demand"].items() if float(v) > 0}
        impacted = False
        impacted_hours = 0
        impacted_resources = set()

        for h in range(int(task["base_start"]), int(task["base_finish"]) + 1):
            if h not in cap.index:
                continue

            for r, demand in demanded.items():
                avail = float(cap.at[h, r])

                if avail + 1e-9 < float(demand):
                    impacted = True
                    impacted_hours += 1
                    impacted_resources.add(r)

        rows.append({
            "task_id": tid,
            "full_name": task["full_name"],
            "Task colpita?": "SI" if impacted else "NO",
            "Ore potenzialmente critiche": impacted_hours,
            "Risorse critiche": ", ".join(sorted(impacted_resources)),
            "Criticality": task["criticality"]
        })

    out = pd.DataFrame(rows)
    out = out.sort_values(
        ["Task colpita?", "Ore potenzialmente critiche", "Criticality"],
        ascending=[False, False, False]
    ).reset_index(drop=True)

    return out


def build_successor_map(tasks):
    succ = defaultdict(list)

    for tid, task in tasks.items():
        for p in task["preds"]:
            succ[p].append(tid)

    return succ


def get_indirectly_impacted_tasks(tasks, direct_impacted_ids):
    succ = build_successor_map(tasks)

    indirect = set()
    frontier = list(direct_impacted_ids)

    while frontier:
        current = frontier.pop()

        for s in succ.get(current, []):
            if s not in direct_impacted_ids and s not in indirect:
                indirect.add(s)
                frontier.append(s)

    return indirect


def schedule_tasks(tasks, capacities_df, allow_early_start=False):
    cap = capacities_df.copy().sort_values("Ora Assoluta").set_index("Ora Assoluta")
    cal_hours = list(cap.index)

    for task in tasks.values():
        task["remaining"] = int(task["effective_duration"])
        task["start"] = None
        task["finish"] = None
        task["assigned_hours"] = []

    hourly_rows = []
    blocked_rows = []

    task_ids_sorted = sorted(
        tasks.keys(),
        key=lambda tid: (-tasks[tid]["criticality"], tasks[tid]["base_start"], tid)
    )

    for hour in cal_hours:
        remaining_cap = cap.loc[hour, RESOURCE_COLS].astype(float).to_dict()

        candidates = []

        for tid in task_ids_sorted:
            task = tasks[tid]

            if task["remaining"] <= 0:
                continue

            preds_done = all(
                tasks[p]["finish"] is not None and tasks[p]["finish"] < hour
                for p in task["preds"]
            )

            if not preds_done:
                continue

            earliest_by_preds = max([tasks[p]["finish"] for p in task["preds"]], default=0) + 1
            earliest = earliest_by_preds if allow_early_start else max(task["base_start"], earliest_by_preds)

            if hour >= earliest:
                candidates.append(tid)

        for tid in candidates:
            task = tasks[tid]
            demand = task["effective_demand"]

            feasible = True
            lacking = []

            for r, req in demand.items():
                if float(req) > float(remaining_cap.get(r, 0.0)) + 1e-9:
                    feasible = False
                    lacking.append(r)

            if feasible:
                for r, req in demand.items():
                    remaining_cap[r] -= float(req)

                task["remaining"] -= 1
                task["assigned_hours"].append(hour)

                if task["start"] is None:
                    task["start"] = hour

                task["finish"] = hour

                hourly_rows.append({
                    "Ora Assoluta": hour,
                    "task_id": tid,
                    "full_name": task["full_name"],
                    **{r: float(demand[r]) for r in RESOURCE_COLS}
                })
            else:
                blocked_rows.append({
                    "Ora Assoluta": hour,
                    "task_id": tid,
                    "full_name": task["full_name"],
                    "Motivo": "Capacità insufficiente",
                    "Risorse mancanti": ", ".join(lacking)
                })

    for tid, task in tasks.items():
        if task["remaining"] > 0:
            blocked_rows.append({
                "Ora Assoluta": np.nan,
                "task_id": tid,
                "full_name": task["full_name"],
                "Motivo": "Task non completata",
                "Risorse mancanti": ""
            })

    return pd.DataFrame(hourly_rows), pd.DataFrame(blocked_rows)


def schedule_tasks_no_recovery_locked(tasks, capacities_df):
    cap = capacities_df.copy().sort_values("Ora Assoluta").set_index("Ora Assoluta")

    for task in tasks.values():
        task["effective_duration"] = int(task["duration"])
        task["effective_demand"] = dict(task["base_demand"])
        task["remaining"] = int(task["duration"])
        task["start"] = None
        task["finish"] = None
        task["assigned_hours"] = []
        task["lost_hours"] = 0

    hourly_rows = []
    blocked_rows = []

    task_ids_sorted = sorted(tasks.keys(), key=lambda tid: (tasks[tid]["base_start"], tid))

    for tid in task_ids_sorted:
        task = tasks[tid]

        pred_finish = []

        for p in task["preds"]:
            if p in tasks and tasks[p]["finish"] is not None:
                pred_finish.append(tasks[p]["finish"])

        earliest_by_preds = max(pred_finish, default=0) + 1
        planned_start = max(task["base_start"], earliest_by_preds)

        current_hour = planned_start
        worked_hours = 0

        while worked_hours < int(task["duration"]):

            if current_hour not in cap.index:
                blocked_rows.append({
                    "Ora Assoluta": current_hour,
                    "task_id": tid,
                    "full_name": task["full_name"],
                    "Motivo": "Ora non presente nel calendario",
                    "Risorse mancanti": ""
                })
                current_hour += 1
                continue

            demand = task["base_demand"]
            feasible = True
            lacking = []

            for r, req in demand.items():
                req = float(req)
                available = float(cap.at[current_hour, r])

                if req > available + 1e-9:
                    feasible = False
                    lacking.append(r)

            if feasible:
                if task["start"] is None:
                    task["start"] = current_hour

                task["assigned_hours"].append(current_hour)
                task["finish"] = current_hour
                worked_hours += 1

                hourly_rows.append({
                    "Ora Assoluta": current_hour,
                    "task_id": tid,
                    "full_name": task["full_name"],
                    **{r: float(demand[r]) for r in RESOURCE_COLS}
                })

            else:
                task["lost_hours"] += 1

                blocked_rows.append({
                    "Ora Assoluta": current_hour,
                    "task_id": tid,
                    "full_name": task["full_name"],
                    "Motivo": "Blocco imprevisto / capacità insufficiente",
                    "Risorse mancanti": ", ".join(lacking)
                })

            current_hour += 1

        task["remaining"] = 0

    return pd.DataFrame(hourly_rows), pd.DataFrame(blocked_rows)


def build_pl_rescheduled(pl_clean, tasks, calendar_df):
    out = pl_clean.copy()

    out["Nuova Ora assoluta Inizio"] = out["task_id"].map({tid: t["start"] for tid, t in tasks.items()})
    out["Nuova Ora assoluta Fine"] = out["task_id"].map({tid: t["finish"] for tid, t in tasks.items()})
    out["Durata effettiva (h)"] = out["task_id"].map({tid: t["effective_duration"] for tid, t in tasks.items()})
    out["Moltiplicatore recovery"] = out["task_id"].map({tid: t.get("recovery_multiplier", t["baseline_multiplier"]) for tid, t in tasks.items()})

    out["Ritardo fine (h)"] = (
        pd.to_numeric(out["Nuova Ora assoluta Fine"], errors="coerce")
        - pd.to_numeric(out["Ora assoluta Fine"], errors="coerce")
    ).fillna(0).astype(int)

    out["Ritardo fine (gg lav)"] = (out["Ritardo fine (h)"] / ORE_GIORNO).round(2)

    cal_map = calendar_df.set_index("Ora Assoluta")[["Data", "Ora"]].to_dict("index")

    def map_date(abs_hour):
        if pd.isna(abs_hour):
            return pd.NaT
        return cal_map.get(int(abs_hour), {}).get("Data", pd.NaT)

    def map_hour(abs_hour):
        if pd.isna(abs_hour):
            return None
        return cal_map.get(int(abs_hour), {}).get("Ora", None)

    out["Nuova Data Inizio"] = out["Nuova Ora assoluta Inizio"].apply(map_date)
    out["Nuova Ora Inizio"] = out["Nuova Ora assoluta Inizio"].apply(map_hour)
    out["Nuova Data Fine"] = out["Nuova Ora assoluta Fine"].apply(map_date)
    out["Nuova Ora Fine"] = out["Nuova Ora assoluta Fine"].apply(map_hour)

    return out


def build_orario_rescheduled(orario_df, hourly_df, id_col):
    out = orario_df.copy()

    if hourly_df.empty:
        out[id_col] = ""
        return out

    alloc = (
        hourly_df.groupby("Ora Assoluta")
        .agg({
            "task_id": lambda s: ";".join(str(int(x)) for x in sorted(s))
        })
        .reset_index()
    )

    alloc_map = dict(zip(alloc["Ora Assoluta"], alloc["task_id"]))
    out[id_col] = out["Ora Assoluta"].map(alloc_map).fillna("")

    return out


def build_delayed_tasks_report(pl_res):
    cols = [
        "task_id",
        "Elemento",
        "Nome attività",
        "full_name",
        "Ora assoluta Inizio",
        "Ora assoluta Fine",
        "Nuova Ora assoluta Inizio",
        "Nuova Ora assoluta Fine",
        "Ritardo fine (h)",
        "Ritardo fine (gg lav)",
        "Predecessori"
    ]

    out = pl_res[cols].copy()
    out = out[out["Ritardo fine (h)"] > 0].copy()
    out = out.sort_values(["Ritardo fine (h)", "Nuova Ora assoluta Fine"], ascending=[False, True]).reset_index(drop=True)

    return out


def calculate_recovery_candidate(task, impact_type):
    if task["name"] not in RECOVERABLE_ACTIVITIES:
        return None

    if int(task["duration"]) < MIN_TASK_DURATION_FOR_COMPRESSION:
        return None

    baseline_multiplier = float(task["baseline_multiplier"])
    recovery_multiplier = baseline_multiplier + MAX_EXTRA_TEAMS_PER_TASK
    extra_teams = recovery_multiplier - baseline_multiplier

    base_duration = int(task["duration"])

    proposed_duration = int(math.ceil(
        base_duration * (baseline_multiplier / recovery_multiplier)
    ))

    min_allowed_duration = int(math.ceil(base_duration * MIN_DURATION_RATIO))
    proposed_duration = max(proposed_duration, min_allowed_duration)

    if proposed_duration >= base_duration:
        return None

    new_demand = {}

    ratio = recovery_multiplier / baseline_multiplier

    # Operai: scalano sempre col moltiplicatore
    for r in LABOR_COLS:
        new_demand[r] = round(float(task["base_demand"][r]) * ratio, 2)

    # Macchinari: scalano secondo regola attività
    machine_rule = MACHINE_SCALE_RULES.get(task["name"], "never")

    for r in MACHINE_COLS:
        base_machine_qty = float(task["base_demand"].get(r, 0.0))

        if not task["machine_flags"].get(r, False):
            new_demand[r] = 0.0
            continue

        if machine_rule == "always":
            # Getto / Posa:
            # la macchina scala subito col moltiplicatore
            new_demand[r] = recovery_multiplier

        elif machine_rule == "after_first_extra_team":
            # Armatura / Casseratura:
            # primo aumento: solo operai
            # dal secondo aumento in poi: anche macchine
            if extra_teams <= 1:
                new_demand[r] = base_machine_qty
            else:
                new_demand[r] = recovery_multiplier

        else:
            # never
            new_demand[r] = base_machine_qty

    extra_demand = {
        r: max(0.0, float(new_demand[r]) - float(task["base_demand"][r]))
        for r in RESOURCE_COLS
    }

    recovery_hours = proposed_duration
    total_extra_cost = 0.0
    extra_cost_cols = {}

    # Costi operai
    for r in LABOR_COLS:
        extra_qty = float(extra_demand.get(r, 0.0))
        cost_r = extra_qty * COSTI_OPERAI_ORARI.get(r, 0) * recovery_hours
        extra_cost_cols[f"Costo extra {r}"] = cost_r
        total_extra_cost += cost_r

    # Costi macchinari
    for r in MACHINE_COLS:
        extra_qty = float(extra_demand.get(r, 0.0))

        if extra_qty <= 0:
            cost_r = 0.0
        else:
            cfg = COSTI_MACCHINARI.get(r, {"tipo_costo": "orario", "costo": 0})

            if cfg["tipo_costo"] == "giornaliero":
                giorni = math.ceil(recovery_hours / ORE_GIORNO)
                cost_r = extra_qty * cfg["costo"] * giorni
            else:
                cost_r = extra_qty * cfg["costo"] * recovery_hours

        extra_cost_cols[f"Costo extra {r}"] = cost_r
        total_extra_cost += cost_r

    base_cost = float(task["base_cost"])
    recovered_hours = base_duration - proposed_duration

    return {
        "task_id": task["id"],
        "full_name": task["full_name"],
        "Tipo impatto recovery": impact_type,
        "Regola macchinari": machine_rule,

        "Baseline multiplier": baseline_multiplier,
        "Recovery multiplier": recovery_multiplier,
        "Extra teams": extra_teams,

        "Durata baseline": base_duration,
        "Durata recovery": proposed_duration,
        "Ore recovery": proposed_duration,
        "Ore recuperate": recovered_hours,

        "Costo attività": base_cost,

        **{f"Extra {r}": extra_demand[r] for r in RESOURCE_COLS},
        **extra_cost_cols,

        "Costo extra totale": total_extra_cost,
        "% costo extra totale": total_extra_cost / base_cost if base_cost > 0 else 0.0,
        "Costo per ora recuperata": total_extra_cost / recovered_hours if recovered_hours > 0 else np.inf,

        "Criticality": task["criticality"],

        "_effective_duration": proposed_duration,
        "_effective_demand": new_demand,
        "_recovery_multiplier": recovery_multiplier,
    }

def apply_recovery_to_task(tasks, candidate_row):
    tid = int(candidate_row["task_id"])
    task = tasks[tid]

    task["effective_duration"] = int(candidate_row["_effective_duration"])
    task["effective_demand"] = dict(candidate_row["_effective_demand"])
    task["recovery_multiplier"] = float(candidate_row["_recovery_multiplier"])


def propose_recovery_plan(tasks, task_colpite_df, imprevisti_df, target_recovery_hours=None):
    direct_ids = set(
        task_colpite_df.loc[
            task_colpite_df["Task colpita?"] == "SI",
            "task_id"
        ].tolist()
    )

    candidates = []

    for tid in direct_ids:
        cand = calculate_recovery_candidate(tasks[tid], "Diretta")
        if cand is not None:
            candidates.append(cand)

    candidates_df = pd.DataFrame(candidates)

    if candidates_df.empty:
        return pd.DataFrame()

    candidates_df = candidates_df.sort_values(
        ["Criticality", "Ore recuperate", "Costo per ora recuperata"],
        ascending=[False, False, True]
    ).reset_index(drop=True)

    plan_rows = []
    recovered_total = 0

    for _, row in candidates_df.iterrows():
        apply_recovery_to_task(tasks, row)
        plan_rows.append(row.drop(labels=[c for c in row.index if str(c).startswith("_")]).to_dict())
        recovered_total += float(row["Ore recuperate"])

        if target_recovery_hours is not None and recovered_total >= target_recovery_hours:
            break

    return pd.DataFrame(plan_rows)


def propose_recovery_plan_economic_chain(tasks, task_colpite_df, imprevisti_df, target_recovery_hours=None):
    direct_ids = set(
        task_colpite_df.loc[
            task_colpite_df["Task colpita?"] == "SI",
            "task_id"
        ].tolist()
    )

    indirect_ids = get_indirectly_impacted_tasks(tasks, direct_ids)
    all_ids = list(direct_ids.union(indirect_ids))

    candidates = []

    for tid in all_ids:
        impact_type = "Diretta" if tid in direct_ids else "Indiretta"
        cand = calculate_recovery_candidate(tasks[tid], impact_type)

        if cand is not None:
            candidates.append(cand)

    candidates_df = pd.DataFrame(candidates)

    if candidates_df.empty:
        return pd.DataFrame(), candidates_df

    candidates_df = candidates_df.sort_values(
        ["Costo per ora recuperata", "Criticality", "Ore recuperate"],
        ascending=[True, False, False]
    ).reset_index(drop=True)

    plan_rows = []
    recovered_total = 0

    for _, row in candidates_df.iterrows():
        apply_recovery_to_task(tasks, row)
        plan_rows.append(row.drop(labels=[c for c in row.index if str(c).startswith("_")]).to_dict())
        recovered_total += float(row["Ore recuperate"])

        if target_recovery_hours is not None and recovered_total >= target_recovery_hours:
            break

    return pd.DataFrame(plan_rows), candidates_df.drop(columns=[c for c in candidates_df.columns if str(c).startswith("_")])


def style_output_excel(filepath):
    wb = load_workbook(filepath)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter

            for cell in col_cells:
                try:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                except:
                    pass

            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

        ws.freeze_panes = "A2"

    wb.save(filepath)



def load_input_workbook(input_xlsx):
    input_path = Path(input_xlsx).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"Excel file not found: {input_path}")

    xls = pd.ExcelFile(input_path)
    required_sheets = {"PL", "Orario"}
    missing = required_sheets.difference(set(xls.sheet_names))
    if missing:
        raise ValueError(f"Nel file Excel mancano i fogli: {sorted(missing)}. Fogli presenti: {xls.sheet_names}")

    pl = pd.read_excel(xls, "PL")
    orario = pd.read_excel(xls, "Orario")
    pl.columns = [str(c).strip() for c in pl.columns]
    orario.columns = [str(c).strip() for c in orario.columns]

    id_column = find_column(pl.columns, "WBS")
    orario_id_column = id_column if id_column in orario.columns else find_column(orario.columns, "WBS")
    dur_column = find_column(pl.columns, "h arrot")

    pl["task_id"] = pd.to_numeric(pl[id_column], errors="coerce")
    pl = pl.dropna(subset=["task_id"]).copy()
    pl["task_id"] = pl["task_id"].astype(int)

    pl["duration_hours"] = pd.to_numeric(pl[dur_column], errors="coerce").fillna(0).astype(int)

    for col in ["Elemento", "Nome attività"]:
        if col not in pl.columns:
            raise ValueError(f"Manca la colonna '{col}' nel foglio PL.")

    pl["Elemento"] = pl["Elemento"].astype(str).str.strip()
    pl["Nome attività"] = pl["Nome attività"].astype(str).str.strip()
    pl["full_name"] = pl["Elemento"] + " - " + pl["Nome attività"]

    if "Costo" not in pl.columns:
        print("Attenzione: colonna 'Costo' non trovata nel PL. Imposto Costo = 0.")
        pl["Costo"] = 0.0

    return pl, orario, id_column, orario_id_column


def run_simulation(input_xlsx, output_xlsx, no_prompt=False, terminal_prompt=False):
    global LABOR_COLS, MACHINE_COLS, RESOURCE_COLS

    pl_df, orario_df, id_col, orario_id_col = load_input_workbook(input_xlsx)
    LABOR_COLS, MACHINE_COLS, RESOURCE_COLS = infer_resource_columns(pl_df)

    print("Input caricati.")
    print("PL:", pl_df.shape)
    print("Orario:", orario_df.shape)
    print("Colonna ID attività PL:", id_col)
    print("Colonna ID attività Orario:", orario_id_col)
    print("Risorse operative:", RESOURCE_COLS)
    show_df(pl_df.head(), "Anteprima PL")

    available_hours = sorted(orario_df["Ora"].astype(str).map(normalize_hour_label).unique().tolist())
    available_dates = (
        pd.to_datetime(orario_df["Data"], errors="coerce")
        .dropna()
        .dt.strftime("%Y-%m-%d")
        .drop_duplicates()
        .sort_values()
        .tolist()
    )

    imprevisti_rows = list(IMPREVISTI_INPUT)
    if not imprevisti_rows and not no_prompt:
        if terminal_prompt:
            imprevisti_rows = prompt_imprevisti(available_hours)
        else:
            imprevisti_rows = prompt_imprevisti_gui(available_hours, available_dates)

    imprevisti_df = build_imprevisti_df(imprevisti_rows)
    if imprevisti_df.empty:
        raise ValueError(
            "Nessun imprevisto inserito. Compila IMPREVISTI_INPUT nello script oppure esegui senza --no-prompt."
        )

    pl_clean, tasks = build_tasks(pl_df)
    calendar_df = build_extended_calendar(orario_df, extra_business_days=EXTRA_BUSINESS_DAYS)
    peak_caps = compute_original_peak_capacities(orario_df, pl_clean, orario_id_col)

    show_df(imprevisti_df, "IMPREVISTI INPUT")
    imprevisti_clean = preprocess_imprevisti(imprevisti_df)

    capacities_df, imprevisti_log_df = build_capacity_table_from_imprevisti(
        calendar_df=calendar_df,
        peak_caps=peak_caps,
        imprevisti_df=imprevisti_clean,
    )
    show_df(imprevisti_log_df, "LOG IMPREVISTI APPLICATI")

    task_colpite_df = identify_impacted_tasks(tasks, capacities_df)
    attivita_colpite_df = task_colpite_df[task_colpite_df["Task colpita?"] == "SI"].copy()
    show_df(attivita_colpite_df, "ATTIVITÀ DIRETTAMENTE COLPITE DAGLI IMPREVISTI")

    # Scenario 1 - senza recovery
    tasks_no_recovery = clone_tasks(tasks)
    hourly_df, blocked_df = schedule_tasks_no_recovery_locked(tasks=tasks_no_recovery, capacities_df=capacities_df)
    pl_res = build_pl_rescheduled(pl_clean, tasks_no_recovery, calendar_df)
    orario_res = build_orario_rescheduled(orario_df, hourly_df, orario_id_col)
    delayed_tasks_df = build_delayed_tasks_report(pl_res)

    baseline_finish = int(pl_clean["Ora assoluta Fine"].max())
    finish_no_recovery = int(pl_res["Nuova Ora assoluta Fine"].max())
    delay_no_recovery = finish_no_recovery - baseline_finish
    enable_recovery_run = bool(ENABLE_RECOVERY and delay_no_recovery > 0)

    # Default output recovery
    recovery_plan_df = pd.DataFrame()
    capacities_recovery_df = pd.DataFrame()
    recovery_log_df = pd.DataFrame()
    hourly_recovery_df = pd.DataFrame()
    blocked_recovery_df = pd.DataFrame()
    pl_recovery = pd.DataFrame()
    delayed_tasks_recovery_df = pd.DataFrame()
    orario_recovery = pd.DataFrame()

    recovery_chain_plan_df = pd.DataFrame()
    recovery_chain_candidates_df = pd.DataFrame()
    hourly_recovery_chain_df = pd.DataFrame()
    blocked_recovery_chain_df = pd.DataFrame()
    pl_recovery_chain = pd.DataFrame()
    delayed_tasks_recovery_chain_df = pd.DataFrame()
    orario_recovery_chain = pd.DataFrame()

    if enable_recovery_run:
        tasks_recovery = clone_tasks(tasks)
        recovery_plan_df = propose_recovery_plan(
            tasks=tasks_recovery,
            task_colpite_df=task_colpite_df,
            imprevisti_df=imprevisti_clean,
            target_recovery_hours=delay_no_recovery,
        )
        capacities_recovery_df = capacities_df.copy()
        recovery_log_df = recovery_plan_df.copy()
        hourly_recovery_df, blocked_recovery_df = schedule_tasks(
            tasks=tasks_recovery,
            capacities_df=capacities_recovery_df,
            allow_early_start=False,
        )
        pl_recovery = build_pl_rescheduled(pl_clean, tasks_recovery, calendar_df)
        delayed_tasks_recovery_df = build_delayed_tasks_report(pl_recovery)
        orario_recovery = build_orario_rescheduled(orario_df, hourly_recovery_df, orario_id_col)

        tasks_recovery_chain = clone_tasks(tasks)
        recovery_chain_plan_df, recovery_chain_candidates_df = propose_recovery_plan_economic_chain(
            tasks=tasks_recovery_chain,
            task_colpite_df=task_colpite_df,
            imprevisti_df=imprevisti_clean,
            target_recovery_hours=delay_no_recovery,
        )
        hourly_recovery_chain_df, blocked_recovery_chain_df = schedule_tasks(
            tasks=tasks_recovery_chain,
            capacities_df=capacities_df.copy(),
            allow_early_start=False,
        )
        pl_recovery_chain = build_pl_rescheduled(pl_clean, tasks_recovery_chain, calendar_df)
        delayed_tasks_recovery_chain_df = build_delayed_tasks_report(pl_recovery_chain)
        orario_recovery_chain = build_orario_rescheduled(orario_df, hourly_recovery_chain_df, orario_id_col)

    comparison_rows = [
        {
            "Scenario": "Baseline",
            "Fine progetto (ora assoluta)": baseline_finish,
            "Ritardo vs baseline (h)": 0,
            "Ritardo vs baseline (gg lav)": 0.0,
        },
        {
            "Scenario": "Con imprevisti - senza recovery",
            "Fine progetto (ora assoluta)": finish_no_recovery,
            "Ritardo vs baseline (h)": finish_no_recovery - baseline_finish,
            "Ritardo vs baseline (gg lav)": round((finish_no_recovery - baseline_finish) / ORE_GIORNO, 2),
        },
    ]

    if enable_recovery_run and not pl_recovery.empty:
        finish_recovery = int(pl_recovery["Nuova Ora assoluta Fine"].max())
        comparison_rows.append(
            {
                "Scenario": "Con imprevisti + recovery diretta",
                "Fine progetto (ora assoluta)": finish_recovery,
                "Ritardo vs baseline (h)": finish_recovery - baseline_finish,
                "Ritardo vs baseline (gg lav)": round((finish_recovery - baseline_finish) / ORE_GIORNO, 2),
            }
        )

    if enable_recovery_run and not pl_recovery_chain.empty:
        finish_recovery_chain = int(pl_recovery_chain["Nuova Ora assoluta Fine"].max())
        comparison_rows.append(
            {
                "Scenario": "Con imprevisti + recovery economica catena",
                "Fine progetto (ora assoluta)": finish_recovery_chain,
                "Ritardo vs baseline (h)": finish_recovery_chain - baseline_finish,
                "Ritardo vs baseline (gg lav)": round((finish_recovery_chain - baseline_finish) / ORE_GIORNO, 2),
            }
        )

    comparison_df = pd.DataFrame(comparison_rows)
    show_df(comparison_df, "CONFRONTO SCENARI")

    scenario_df = pd.DataFrame([
        {
            "Scenario": "Imprevisti applicati",
            "Recovery attivo": "SI" if ENABLE_RECOVERY else "NO",
            "Recovery eseguita": "SI" if enable_recovery_run else "NO",
            "Recovery logica": "Moltiplicatore squadra +1 max",
            "Max extra team per task": MAX_EXTRA_TEAMS_PER_TASK,
            "Min duration ratio": MIN_DURATION_RATIO,
        }
    ])

    costo_totale_progetto = float(pl_clean["Costo"].sum()) if "Costo" in pl_clean.columns else 0.0
    cost_summary_rows = []

    if enable_recovery_run and not recovery_plan_df.empty:
        costo_extra_diretto = float(recovery_plan_df["Costo extra totale"].sum())
        costo_attivita_dirette = float(recovery_plan_df["Costo attività"].sum())
        cost_summary_rows.append({
            "Scenario": "Recovery diretta",
            "Costo totale baseline progetto": costo_totale_progetto,
            "Costo attività coinvolte": costo_attivita_dirette,
            "Costo extra totale recovery": costo_extra_diretto,
            "Costo totale progetto con recovery": costo_totale_progetto + costo_extra_diretto,
            "% extra su attività coinvolte": costo_extra_diretto / costo_attivita_dirette if costo_attivita_dirette > 0 else 0,
            "% extra su totale progetto": costo_extra_diretto / costo_totale_progetto if costo_totale_progetto > 0 else 0,
        })

    if enable_recovery_run and not recovery_chain_plan_df.empty:
        costo_extra_chain = float(recovery_chain_plan_df["Costo extra totale"].sum())
        costo_attivita_chain = float(recovery_chain_plan_df["Costo attività"].sum())
        cost_summary_rows.append({
            "Scenario": "Recovery economica catena",
            "Costo totale baseline progetto": costo_totale_progetto,
            "Costo attività coinvolte": costo_attivita_chain,
            "Costo extra totale recovery": costo_extra_chain,
            "Costo totale progetto con recovery": costo_totale_progetto + costo_extra_chain,
            "% extra su attività coinvolte": costo_extra_chain / costo_attivita_chain if costo_attivita_chain > 0 else 0,
            "% extra su totale progetto": costo_extra_chain / costo_totale_progetto if costo_totale_progetto > 0 else 0,
        })

    cost_summary_df = pd.DataFrame(cost_summary_rows)

    output_path = Path(output_xlsx).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pl_df.to_excel(writer, sheet_name="PL_originale", index=False)
        orario_df.to_excel(writer, sheet_name="Orario_originale", index=False)
        imprevisti_clean.to_excel(writer, sheet_name="Imprevisti_letti", index=False)
        capacities_df.to_excel(writer, sheet_name="Capacita_imprevisti", index=False)
        imprevisti_log_df.to_excel(writer, sheet_name="Log_imprevisti", index=False)
        task_colpite_df.to_excel(writer, sheet_name="Task_colpite", index=False)
        attivita_colpite_df.to_excel(writer, sheet_name="Attivita_colpite", index=False)
        blocked_df.to_excel(writer, sheet_name="Task_bloccate", index=False)
        hourly_df.to_excel(writer, sheet_name="Dettaglio_allocazione", index=False)
        pl_res.to_excel(writer, sheet_name="PL_riprogrammato", index=False)
        orario_res.to_excel(writer, sheet_name="Orario_riprogrammato", index=False)
        delayed_tasks_df.to_excel(writer, sheet_name="Report_ritardi", index=False)
        scenario_df.to_excel(writer, sheet_name="Scenario", index=False)
        comparison_df.to_excel(writer, sheet_name="Confronto_scenari", index=False)
        cost_summary_df.to_excel(writer, sheet_name="Costo_recovery", index=False)

        if enable_recovery_run:
            recovery_plan_df.to_excel(writer, sheet_name="Proposta_recovery_dir", index=False)
            capacities_recovery_df.to_excel(writer, sheet_name="Capacita_recovery_dir", index=False)
            recovery_log_df.to_excel(writer, sheet_name="Recovery_log_dir", index=False)
            hourly_recovery_df.to_excel(writer, sheet_name="Dettaglio_recovery_dir", index=False)
            blocked_recovery_df.to_excel(writer, sheet_name="Task_bloccate_dir", index=False)
            pl_recovery.to_excel(writer, sheet_name="PL_recovery_dir", index=False)
            orario_recovery.to_excel(writer, sheet_name="Orario_recovery_dir", index=False)
            delayed_tasks_recovery_df.to_excel(writer, sheet_name="Report_recovery_dir", index=False)
            recovery_chain_candidates_df.to_excel(writer, sheet_name="Candidate_chain", index=False)
            recovery_chain_plan_df.to_excel(writer, sheet_name="Proposta_chain", index=False)
            hourly_recovery_chain_df.to_excel(writer, sheet_name="Dettaglio_chain", index=False)
            blocked_recovery_chain_df.to_excel(writer, sheet_name="Task_bloccate_chain", index=False)
            pl_recovery_chain.to_excel(writer, sheet_name="PL_recovery_chain", index=False)
            orario_recovery_chain.to_excel(writer, sheet_name="Orario_chain", index=False)
            delayed_tasks_recovery_chain_df.to_excel(writer, sheet_name="Report_chain", index=False)

    style_output_excel(output_path)
    print(f"\nFile creato: {output_path}")

    if cost_summary_df.empty:
        print("Nessun costo recovery da stampare.")
    else:
        cost_print = cost_summary_df.copy()
        money_cols = [
            "Costo totale baseline progetto",
            "Costo attività coinvolte",
            "Costo extra totale recovery",
            "Costo totale progetto con recovery",
        ]
        pct_cols = ["% extra su attività coinvolte", "% extra su totale progetto"]
        for c in money_cols:
            cost_print[c] = cost_print[c].apply(euro_it)
        for c in pct_cols:
            cost_print[c] = cost_print[c].apply(pct_it)
        show_df(cost_print, "COSTI RECOVERY - RIEPILOGO")

    return output_path




def launch_full_gui(default_input=INPUT_XLSX, default_output=OUTPUT_XLSX):
    """
    Clean GUI v10: no tkcalendar dependency; stable built-in date picker for Windows and macOS.
    Results are shown in tabs with readable tables; technical logs are separated.
    """
    import os
    import sys
    import io
    import threading
    import traceback
    import subprocess
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    from pathlib import Path

    # Date picker.
    # On Windows/Linux we use tkcalendar.DateEntry, which is the exact old calendar
    # style requested by the user. On macOS we keep the built-in Tk fallback because
    # tkcalendar can freeze on some macOS/Tk builds when opening the popup.
    import datetime as _dt
    import calendar as _cal

    DateEntry = None
    if sys.platform != "darwin":
        try:
            from tkcalendar import DateEntry as DateEntry
        except Exception:
            DateEntry = None

    if DateEntry is None:
        class DateEntry(ttk.Frame):
            def __init__(self, master=None, width=16, date_pattern="yyyy-mm-dd", state="readonly", **kwargs):
                super().__init__(master, **kwargs)
                self._date = _dt.date.today()
                self.var = tk.StringVar(value=self._date.isoformat())
                # Use a readonly Combobox only as the visible control, so the arrow
                # matches the other dropdowns in the interface. The click still opens
                # the custom Tk calendar instead of a normal dropdown list.
                self.combo = ttk.Combobox(self, textvariable=self.var, values=[self._date.isoformat()], state="readonly", width=width)
                self.combo.pack(side="left", fill="x", expand=True)
                self.combo.bind("<Button-1>", self._on_open_calendar, add="+")
                self.combo.bind("<Return>", self._on_open_calendar, add="+")
                self.combo.bind("<space>", self._on_open_calendar, add="+")
                self.combo.bind("<<ComboboxSelected>>", self._on_open_calendar, add="+")

            def _on_open_calendar(self, event=None):
                self._open_calendar()
                return "break"

            def set_date(self, value):
                try:
                    if value is None or pd.isna(value):
                        value = _dt.date.today()
                    elif isinstance(value, _dt.datetime):
                        value = value.date()
                    elif isinstance(value, pd.Timestamp):
                        value = value.date()
                    elif isinstance(value, str):
                        value = pd.to_datetime(value, errors="coerce")
                        if pd.isna(value):
                            value = _dt.date.today()
                        else:
                            value = value.date()
                    elif not isinstance(value, _dt.date):
                        value = pd.to_datetime(value, errors="coerce")
                        if pd.isna(value):
                            value = _dt.date.today()
                        else:
                            value = value.date()
                except Exception:
                    value = _dt.date.today()
                self._date = value
                self.var.set(value.isoformat())

            def get_date(self):
                try:
                    return pd.to_datetime(self.var.get(), errors="raise").date()
                except Exception:
                    return self._date

            def _open_calendar(self):
                """Open a built-in calendar popup with the same visual style as the old calendar."""
                current = self.get_date()
                popup = tk.Toplevel(self)
                popup.title("Select date")
                popup.transient(self.winfo_toplevel())
                popup.resizable(False, False)
                popup.configure(bg="#3b3f42")
                popup.attributes("-topmost", True)

                try:
                    x = self.winfo_rootx()
                    y = self.winfo_rooty() + self.winfo_height() + 4
                    popup.geometry(f"+{x}+{y}")
                except Exception:
                    pass

                year_var = tk.IntVar(value=current.year)
                month_var = tk.IntVar(value=current.month)

                # Colors chosen to reproduce the older dark-header calendar style.
                bg_dark = "#3b3f42"
                bg_header = "#3b3f42"
                fg_header = "#ffffff"
                bg_week = "#d7d7d7"
                bg_day = "#f2f2f2"
                bg_other = "#e6e6e6"
                fg_day = "#111111"
                fg_other = "#7a7a7a"
                bg_selected = "#4f6f86"
                fg_selected = "#ffffff"
                border = "#9a9a9a"

                outer = tk.Frame(popup, bg=bg_dark, bd=1, relief="solid")
                outer.pack(fill="both", expand=True)

                header = tk.Frame(outer, bg=bg_header)
                header.pack(fill="x")
                body = tk.Frame(outer, bg=border)
                body.pack(fill="both", expand=True)

                title_month = tk.StringVar()
                title_year = tk.StringVar()

                def change_month(delta):
                    m = month_var.get() + delta
                    y = year_var.get()
                    if m < 1:
                        m = 12
                        y -= 1
                    elif m > 12:
                        m = 1
                        y += 1
                    month_var.set(m)
                    year_var.set(y)
                    render()

                def change_year(delta):
                    year_var.set(year_var.get() + delta)
                    render()

                def header_button(parent, text, command):
                    return tk.Button(
                        parent,
                        text=text,
                        command=command,
                        width=3,
                        bd=0,
                        relief="flat",
                        bg=bg_header,
                        fg=fg_header,
                        activebackground="#55595d",
                        activeforeground=fg_header,
                        font=("TkDefaultFont", 9, "bold"),
                        cursor="hand2",
                    )

                # Header layout like the old popup: arrows around month and year.
                header_button(header, "◂", lambda: change_month(-1)).grid(row=0, column=0, padx=(4, 0), pady=4)
                tk.Label(header, textvariable=title_month, bg=bg_header, fg=fg_header, width=12, anchor="center", font=("TkDefaultFont", 10, "bold")).grid(row=0, column=1, pady=4)
                header_button(header, "▸", lambda: change_month(1)).grid(row=0, column=2, padx=(0, 8), pady=4)
                header_button(header, "◂", lambda: change_year(-1)).grid(row=0, column=3, padx=(8, 0), pady=4)
                tk.Label(header, textvariable=title_year, bg=bg_header, fg=fg_header, width=8, anchor="center", font=("TkDefaultFont", 10, "bold")).grid(row=0, column=4, pady=4)
                header_button(header, "▸", lambda: change_year(1)).grid(row=0, column=5, padx=(0, 4), pady=4)

                for i in range(6):
                    header.columnconfigure(i, weight=1)

                def choose(day):
                    self.set_date(_dt.date(year_var.get(), month_var.get(), day))
                    popup.destroy()

                def make_cell(row, col, text, bg, fg, command=None, bold=False):
                    if command is None:
                        widget = tk.Label(
                            body,
                            text=text,
                            bg=bg,
                            fg=fg,
                            width=4,
                            height=1,
                            anchor="center",
                            font=("TkDefaultFont", 9, "bold" if bold else "normal"),
                            bd=0,
                        )
                    else:
                        widget = tk.Button(
                            body,
                            text=text,
                            command=command,
                            bg=bg,
                            fg=fg,
                            width=4,
                            height=1,
                            relief="flat",
                            bd=0,
                            activebackground=bg_selected,
                            activeforeground=fg_selected,
                            font=("TkDefaultFont", 9, "bold" if bold else "normal"),
                            cursor="hand2",
                        )
                    widget.grid(row=row, column=col, padx=1, pady=1, sticky="nsew")
                    return widget

                def render():
                    for child in body.winfo_children():
                        child.destroy()

                    y = year_var.get()
                    m = month_var.get()
                    title_month.set(_cal.month_name[m])
                    title_year.set(str(y))

                    weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
                    make_cell(0, 0, "", bg_header, fg_header, bold=True)
                    for col, name in enumerate(weekdays, start=1):
                        make_cell(0, col, name, bg_week, fg_day, bold=True)

                    month_cal = _cal.Calendar(firstweekday=0).monthdatescalendar(y, m)
                    selected = self.get_date()
                    today = _dt.date.today()

                    for r, week in enumerate(month_cal, start=1):
                        make_cell(r, 0, str(r), bg_header, fg_header, bold=True)
                        for c, day_date in enumerate(week, start=1):
                            in_month = day_date.month == m
                            is_selected = day_date == selected
                            is_today = day_date == today
                            bg = bg_selected if is_selected else (bg_day if in_month else bg_other)
                            fg = fg_selected if is_selected else (fg_day if in_month else fg_other)
                            txt = str(day_date.day)
                            make_cell(
                                r,
                                c,
                                txt,
                                bg,
                                fg,
                                command=(lambda d=day_date: (self.set_date(d), popup.destroy())),
                                bold=is_today or is_selected,
                            )

                    for col in range(8):
                        body.columnconfigure(col, weight=1)

                render()
                popup.lift()
                popup.focus_force()

                # Close when focus leaves the popup, but do it gently to avoid macOS/Tk lockups.
                def close_if_focus_out(_event=None):
                    try:
                        focused = popup.focus_get()
                        if focused is None or not str(focused).startswith(str(popup)):
                            popup.after(150, lambda: popup.destroy() if popup.winfo_exists() else None)
                    except Exception:
                        pass
                popup.bind("<Escape>", lambda _e: popup.destroy())
    # -----------------------------
    # Helpers
    # -----------------------------
    def resolve_input_excel_path(path_text):
        p = Path(path_text).expanduser()
        candidates = []
        if p.is_absolute():
            candidates.append(p)
        else:
            candidates.extend([
                Path.cwd() / p,
                Path(sys.executable).resolve().parent / p if getattr(sys, "frozen", False) else Path(__file__).resolve().parent / p,
                Path(sys.executable).resolve().parent.parent / p if getattr(sys, "frozen", False) else Path(__file__).resolve().parent.parent / p,
            ])
        for candidate in candidates:
            if candidate.exists():
                return candidate.resolve()
        return p.resolve() if not p.is_absolute() else p

    def safe_open_file(path):
        path = Path(path)
        if not path.exists():
            messagebox.showerror("File not found", f"The file does not exist:\n{path}")
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(path))
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(path)])
            else:
                subprocess.Popen(["xdg-open", str(path)])
        except Exception as exc:
            messagebox.showerror("Open file error", str(exc))

    def clear_tree(tree):
        tree.delete(*tree.get_children())

    def format_euro_value(value):
        try:
            if pd.isna(value):
                return ""
            return f"€ {float(value):,.2f}"
        except Exception:
            return value

    def format_percent_value(value):
        try:
            if pd.isna(value):
                return ""
            return f"{float(value) * 100:.2f}%"
        except Exception:
            return value

    def prepare_display_dataframe(df, lang="EN"):
        if df is None or getattr(df, "empty", False):
            return df
        out = df.copy()
        if lang == "EN":
            out = out.rename(columns={
                "task_id": "Task ID",
                "full_name": "Activity",
                "Tipo impatto recovery": "Recovery impact type",
                "Regola macchinari": "Machinery rule",
                "Durata baseline": "Baseline duration (h)",
                "Durata recovery": "Recovery duration (h)",
                "Ore recuperate": "Recovered hours (h)",
                "Costo attività": "Activity cost",
                "Costo extra totale": "Total extra cost",
                "% costo extra totale": "% extra on activity cost",
                "Costo totale baseline progetto": "Total project cost",
                "% extra su totale progetto": "% extra on total project cost",
                "% total extra cost": "% extra on activity cost",
                "Total project cost": "Total project cost",
                "% extra on total project cost": "% extra on total project cost",
                "Costo per ora recuperata": "Cost per recovered hour",
                "Impacted task?": "Impacted?",
                "Potentially critical hours": "Risk-conflict hours (h)",
                "Criticality": "Criticality score",
                "Extra Operai Liv.1": "Extra Workers Lv.1",
                "Extra Operai Liv.2": "Extra Workers Lv.2",
                "Extra Operai Liv.3": "Extra Workers Lv.3",
                "Extra Gru": "Extra Crane",
                "Extra Pompa Calcestruzzo": "Extra Concrete Pump",
                "Extra Escavatore": "Extra Excavator",
                "Extra Autocarro": "Extra Truck",
            })
        # Format money/percent columns for readability in the GUI.
        # IMPORTANT: percentage columns often contain the word "cost" (e.g. "% total extra cost"),
        # so percentages must be detected BEFORE money columns.
        percent_cols = []
        money_cols = []
        for c in list(out.columns):
            cl = str(c).lower()
            if str(c).startswith("%") or "%" in str(c) or "percent" in cl:
                percent_cols.append(c)
                out[c] = out[c].apply(format_percent_value)
            elif any(k in cl for k in ["cost", "costo"]):
                money_cols.append(c)
                out[c] = out[c].apply(format_euro_value)
            elif out[c].dtype == object:
                out[c] = out[c].apply(lambda x: translate_visible_text(x, lang))
        # Rename only true money columns to explicitly show currency. Percent columns must NOT get (€).
        if lang == "EN":
            rename_money_cols = {c: f"{c} (€)" for c in money_cols if "€" not in str(c)}
            out = out.rename(columns=rename_money_cols)
        return out

    def set_tree_dataframe(tree, df, max_rows=500, lang="EN"):
        df = prepare_display_dataframe(df, lang)
        clear_tree(tree)
        if df is None or getattr(df, "empty", False):
            tree["columns"] = []
            return
        view = df.head(max_rows).copy()
        cols = [str(c) for c in view.columns]
        tree["columns"] = cols
        tree["show"] = "headings"
        for c in cols:
            tree.heading(c, text=c)
            # Moderate auto width based on header and first rows
            samples = [str(c)] + [str(x) for x in view[c].head(20).tolist()]
            width = min(max(max(len(x) for x in samples) * 8 + 18, 90), 260)
            tree.column(c, width=width, anchor="center", stretch=True)
        for _, row in view.iterrows():
            tree.insert("", "end", values=["" if pd.isna(row[c]) else translate_visible_text(row[c], "EN") for c in view.columns])

    def append_text(widget, text):
        widget.configure(state="normal")
        widget.insert("end", text)
        widget.see("end")
        widget.configure(state="disabled")

    def set_text(widget, text):
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("end", text)
        widget.configure(state="disabled")

    def display_row_for_risk(row, lang="EN"):
        if lang == "EN":
            return [
                row.get("ID Rischio", ""),
                translate_visible_text(row.get("Evento", ""), "EN"),
                translate_visible_text(row.get("Tipo impatto", ""), "EN"),
                translate_visible_text(row.get("Risorse colpite", ""), "EN"),
                row.get("Data inizio", ""),
                row.get("Ora inizio", ""),
                row.get("Data fine", ""),
                row.get("Ora fine", ""),
            ]
        return [
            row.get("ID Rischio", ""), row.get("Evento", ""), row.get("Tipo impatto", ""),
            row.get("Risorse colpite", ""), row.get("Data inizio", ""), row.get("Ora inizio", ""),
            row.get("Data fine", ""), row.get("Ora fine", ""),
        ]

    def read_sheet_any(path, names):
        """Read a worksheet by exact or fuzzy name. Handles translated/untranslated output names."""
        try:
            xls = pd.ExcelFile(path)
        except Exception:
            return pd.DataFrame()

        wanted = [str(n) for n in names]

        # Exact match first.
        for name in wanted:
            if name in xls.sheet_names:
                try:
                    return pd.read_excel(xls, sheet_name=name)
                except Exception:
                    pass

        def norm(s):
            return re.sub(r"[^a-z0-9]+", "", str(s).lower())

        norm_map = {norm(s): s for s in xls.sheet_names}
        for name in wanted:
            key = norm(name)
            if key in norm_map:
                try:
                    return pd.read_excel(xls, sheet_name=norm_map[key])
                except Exception:
                    pass

        # Fuzzy contains fallback for PyInstaller/Excel name variants.
        wanted_norm = [norm(n) for n in wanted]
        for sheet in xls.sheet_names:
            s_norm = norm(sheet)
            if any(w and (w in s_norm or s_norm in w) for w in wanted_norm):
                try:
                    return pd.read_excel(xls, sheet_name=sheet)
                except Exception:
                    pass

        return pd.DataFrame()

    def money_value(df, scenario_contains, col_candidates):
        if df is None or df.empty or "Scenario" not in df.columns:
            return None
        row = df[df["Scenario"].astype(str).str.contains(scenario_contains, case=False, na=False)]
        if row.empty:
            return None
        for col in col_candidates:
            if col in row.columns:
                return row.iloc[0][col]
        return None

    # -----------------------------
    # App
    # -----------------------------
    app = tk.Tk()
    app.title("Construction Risk Simulation Tool")
    app.geometry("1280x820")
    app.minsize(1120, 720)

    style = ttk.Style(app)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    imprevisti_rows = []
    available_hours_cache = []
    available_dates_cache = []
    output_created_path = tk.StringVar(value="")
    is_running = tk.BooleanVar(value=False)

    input_var = tk.StringVar(value=str(default_input))
    output_var = tk.StringVar(value=str(default_output))
    output_language_var = tk.StringVar(value="English")
    status_var = tk.StringVar(value="Select an input Excel file and load timetables.")

    # Header
    top = ttk.Frame(app, padding=(12, 10))
    top.pack(fill="x")
    ttk.Label(top, text="Construction Risk Simulation Tool", font=("Segoe UI", 16, "bold")).pack(side="left")
    ttk.Label(top, textvariable=status_var).pack(side="right")

    nb = ttk.Notebook(app)
    nb.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    tab_input = ttk.Frame(nb, padding=10)
    tab_summary = ttk.Frame(nb, padding=10)
    tab_scenario = ttk.Frame(nb, padding=10)
    tab_costs = ttk.Frame(nb, padding=10)
    tab_impacts = ttk.Frame(nb, padding=10)
    tab_recovery_details = ttk.Frame(nb, padding=10)
    tab_extra_resources = ttk.Frame(nb, padding=10)
    tab_chain_candidates = ttk.Frame(nb, padding=10)
    tab_log = ttk.Frame(nb, padding=10)
    nb.add(tab_input, text="Input & risk events")
    nb.add(tab_summary, text="Summary")
    nb.add(tab_scenario, text="Scenario comparison")
    nb.add(tab_costs, text="Recovery costs")
    nb.add(tab_impacts, text="Impacted activities")
    nb.add(tab_recovery_details, text="Recovery details")
    nb.add(tab_extra_resources, text="Extra resources")
    nb.add(tab_chain_candidates, text="Chain candidates")
    nb.add(tab_log, text="Technical log")

    # -----------------------------
    # Input tab
    # -----------------------------
    file_frame = ttk.LabelFrame(tab_input, text="Files", padding=10)
    file_frame.pack(fill="x", pady=(0, 10))

    def browse_input():
        filename = filedialog.askopenfilename(
            title="Select input Excel file",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
        )
        if filename:
            input_var.set(filename)
            p = Path(filename)
            output_var.set(str(p.with_name(p.stem + "_pronto_intervento_output.xlsx")))
            load_calendar_values(show_errors=True)

    def browse_output():
        filename = filedialog.asksaveasfilename(
            title="Save output Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            output_var.set(filename)

    ttk.Label(file_frame, text="Input Excel:").grid(row=0, column=0, sticky="e", padx=(0, 8), pady=4)
    ttk.Entry(file_frame, textvariable=input_var).grid(row=0, column=1, sticky="ew", pady=4)
    ttk.Button(file_frame, text="Browse...", command=browse_input).grid(row=0, column=2, padx=(8, 0), pady=4)
    ttk.Button(file_frame, text="Load timetables from Excel", command=lambda: load_calendar_values(show_errors=True)).grid(row=0, column=3, padx=(8, 0), pady=4)

    ttk.Label(file_frame, text="Output Excel:").grid(row=1, column=0, sticky="e", padx=(0, 8), pady=4)
    ttk.Entry(file_frame, textvariable=output_var).grid(row=1, column=1, sticky="ew", pady=4)
    ttk.Button(file_frame, text="Save as...", command=browse_output).grid(row=1, column=2, padx=(8, 0), pady=4)

    ttk.Label(file_frame, text="Output language:").grid(row=1, column=3, sticky="e", padx=(20, 6), pady=4)
    lang_combo = ttk.Combobox(file_frame, textvariable=output_language_var, values=["English", "Italian"], state="readonly", width=12)
    lang_combo.grid(row=1, column=4, sticky="w", pady=4)
    file_frame.columnconfigure(1, weight=1)

    risk_frame = ttk.LabelFrame(tab_input, text="Risk event input", padding=10)
    risk_frame.pack(fill="x", pady=(0, 10))

    risk_names = list(RISK_LIBRARY.keys())
    risk_var = tk.StringVar(value=risk_names[0] if risk_names else "")
    ttk.Label(risk_frame, text="Risk event:").grid(row=0, column=0, sticky="e", padx=(0, 8), pady=5)
    ttk.Combobox(risk_frame, textvariable=risk_var, values=risk_names, state="readonly", width=28).grid(row=0, column=1, sticky="w", pady=5)

    ttk.Label(risk_frame, text="Start date:").grid(row=1, column=0, sticky="e", padx=(0, 8), pady=5)
    start_date = DateEntry(risk_frame, width=16, date_pattern="yyyy-mm-dd", state="readonly")
    start_date.grid(row=1, column=1, sticky="w", pady=5)

    ttk.Label(risk_frame, text="Start time:").grid(row=1, column=2, sticky="e", padx=(30, 8), pady=5)
    start_hour_var = tk.StringVar(value="")
    start_hour_combo = ttk.Combobox(risk_frame, textvariable=start_hour_var, state="readonly", width=20)
    start_hour_combo.grid(row=1, column=3, sticky="w", pady=5)

    ttk.Label(risk_frame, text="End date:").grid(row=2, column=0, sticky="e", padx=(0, 8), pady=5)
    end_date = DateEntry(risk_frame, width=16, date_pattern="yyyy-mm-dd", state="readonly")
    end_date.grid(row=2, column=1, sticky="w", pady=5)

    ttk.Label(risk_frame, text="End time:").grid(row=2, column=2, sticky="e", padx=(30, 8), pady=5)
    end_hour_var = tk.StringVar(value="")
    end_hour_combo = ttk.Combobox(risk_frame, textvariable=end_hour_var, state="readonly", width=20)
    end_hour_combo.grid(row=2, column=3, sticky="w", pady=5)

    btns = ttk.Frame(risk_frame)
    btns.grid(row=3, column=0, columnspan=4, sticky="w", pady=(10, 0))

    table_frame = ttk.LabelFrame(tab_input, text="Risk events added", padding=6)
    table_frame.pack(fill="both", expand=True)

    risk_cols = ["Risk ID", "Event", "Impact type", "Affected resources", "Start date", "Start time", "End date", "End time"]
    risk_table = ttk.Treeview(table_frame, columns=risk_cols, show="headings", height=7)
    for col in risk_cols:
        risk_table.heading(col, text=col)
        risk_table.column(col, width=145 if col in ["Impact type", "Affected resources"] else 110, anchor="center")
    risk_table.pack(side="left", fill="both", expand=True)
    risk_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=risk_table.yview)
    risk_scroll.pack(side="right", fill="y")
    risk_table.configure(yscrollcommand=risk_scroll.set)

    def refresh_risk_table():
        clear_tree(risk_table)
        lang = "EN" if output_language_var.get() == "English" else "IT"
        for row in imprevisti_rows:
            risk_table.insert("", "end", values=display_row_for_risk(row, lang))

    def load_calendar_values(show_errors=False):
        try:
            input_path = resolve_input_excel_path(input_var.get().strip())
            if not input_path.exists():
                status_var.set("Input Excel file not found. Use Browse to select it.")
                if show_errors:
                    messagebox.showerror("Input file not found", f"Input Excel file not found:\n{input_path}")
                return False

            xls = pd.ExcelFile(input_path)
            if "Orario" not in xls.sheet_names:
                raise ValueError("The input workbook does not contain a sheet named 'Orario'.")
            orario = pd.read_excel(xls, "Orario")
            orario.columns = [str(c).strip() for c in orario.columns]
            if "Ora" not in orario.columns or "Data" not in orario.columns:
                raise ValueError("The 'Orario' sheet must contain columns named 'Data' and 'Ora'.")

            hours = sorted(orario["Ora"].dropna().astype(str).map(normalize_hour_label).unique().tolist())
            dates_series = pd.to_datetime(orario["Data"], errors="coerce").dropna()
            dates = sorted(dates_series.dt.strftime("%Y-%m-%d").drop_duplicates().tolist())
            if not hours:
                raise ValueError("No valid time slots were found in the 'Ora' column.")
            if not dates:
                raise ValueError("No valid dates were found in the 'Data' column.")

            available_hours_cache[:] = hours
            available_dates_cache[:] = dates
            start_hour_combo.configure(values=hours)
            end_hour_combo.configure(values=hours)
            start_hour_var.set(hours[0])
            end_hour_var.set(hours[-1])
            first_date = pd.to_datetime(dates[0]).date()
            start_date.set_date(first_date)
            end_date.set_date(first_date)
            input_var.set(str(input_path))
            if not output_var.get().strip() or Path(output_var.get()).name == OUTPUT_XLSX:
                output_var.set(str(input_path.with_name(input_path.stem + "_pronto_intervento_output.xlsx")))
            status_var.set(f"Timetables loaded: {len(hours)} time slots, {len(dates)} dates.")
            set_text(technical_log_text, f"Timetables loaded from: {input_path}\nTime slots: {', '.join(hours)}\nDates: {dates[0]} to {dates[-1]}\n")
            return True
        except Exception as exc:
            status_var.set("Could not load timetables from Excel.")
            set_text(technical_log_text, f"Timetable loading error:\n{exc}\n")
            if show_errors:
                messagebox.showerror("Timetable loading error", str(exc))
            return False

    def add_risk_event():
        if not start_hour_var.get().strip() or not end_hour_var.get().strip():
            if not load_calendar_values(show_errors=True):
                return
        risk_name = risk_var.get().strip()
        if risk_name not in RISK_LIBRARY:
            messagebox.showerror("Error", "Select a valid risk event.")
            return
        try:
            d1 = start_date.get_date().strftime("%Y-%m-%d")
            d2 = end_date.get_date().strftime("%Y-%m-%d")
        except Exception:
            messagebox.showerror("Error", "Select valid dates.")
            return
        row = RISK_LIBRARY[risk_name].copy()
        row["Data inizio"] = d1
        row["Ora inizio"] = start_hour_var.get().strip()
        row["Data fine"] = d2
        row["Ora fine"] = end_hour_var.get().strip()
        imprevisti_rows.append(row)
        refresh_risk_table()
        status_var.set(f"Risk events added: {len(imprevisti_rows)}")

    def remove_selected_risk():
        selected = risk_table.selection()
        if not selected:
            messagebox.showinfo("No selection", "Select a row to remove.")
            return
        for idx in sorted([risk_table.index(i) for i in selected], reverse=True):
            if 0 <= idx < len(imprevisti_rows):
                imprevisti_rows.pop(idx)
        refresh_risk_table()
        status_var.set(f"Risk events added: {len(imprevisti_rows)}")

    def clear_risks():
        imprevisti_rows.clear()
        refresh_risk_table()
        status_var.set("Risk event list cleared.")

    ttk.Button(btns, text="Add risk event", command=add_risk_event).pack(side="left", padx=(0, 8))
    ttk.Button(btns, text="Remove selected", command=remove_selected_risk).pack(side="left", padx=(0, 8))
    ttk.Button(btns, text="Clear list", command=clear_risks).pack(side="left", padx=(0, 18))

    # -----------------------------
    # Summary tab
    # -----------------------------
    summary_top = ttk.LabelFrame(tab_summary, text="Simulation summary", padding=12)
    summary_top.pack(fill="x", pady=(0, 10))
    summary_vars = {
        "baseline": tk.StringVar(value="—"),
        "risk_finish": tk.StringVar(value="—"),
        "risk_delay": tk.StringVar(value="—"),
        "direct_finish": tk.StringVar(value="—"),
        "chain_finish": tk.StringVar(value="—"),
        "direct_cost": tk.StringVar(value="—"),
        "chain_cost": tk.StringVar(value="—"),
        "output": tk.StringVar(value="—"),
    }
    labels = [
        ("Baseline project finish", "baseline"),
        ("Project finish with risks", "risk_finish"),
        ("Delay caused by risks", "risk_delay"),
        ("Project finish with direct recovery", "direct_finish"),
        ("Project finish with optimized chain recovery", "chain_finish"),
        ("Direct recovery extra cost", "direct_cost"),
        ("Optimized chain recovery extra cost", "chain_cost"),
        ("Output file", "output"),
    ]
    for i, (label, key) in enumerate(labels):
        ttk.Label(summary_top, text=label + ":", font=("Segoe UI", 10, "bold")).grid(row=i//2, column=(i%2)*2, sticky="e", padx=(0, 8), pady=5)
        ttk.Label(summary_top, textvariable=summary_vars[key]).grid(row=i//2, column=(i%2)*2+1, sticky="w", pady=5)
    summary_top.columnconfigure(1, weight=1)
    summary_top.columnconfigure(3, weight=1)

    summary_actions = ttk.Frame(tab_summary)
    summary_actions.pack(fill="x", pady=(0, 10))
    run_btn = ttk.Button(summary_actions, text="Run simulation", command=lambda: run_from_gui())
    run_btn.pack(side="left", padx=(0, 8))
    open_output_btn = ttk.Button(summary_actions, text="Open output Excel", command=lambda: safe_open_file(output_created_path.get() or output_var.get()), state="disabled")
    open_output_btn.pack(side="left", padx=(0, 8))
    progress = ttk.Progressbar(summary_actions, mode="indeterminate", length=220)
    progress.pack(side="left", padx=(10, 0))

    log_clean = tk.Text(tab_summary, height=12, wrap="word", state="disabled")
    log_clean.pack(fill="both", expand=True)

    # -----------------------------
    # Tables tabs
    # -----------------------------
    def make_tree(parent):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True)
        tree = ttk.Treeview(frame, show="headings")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        return tree

    scenario_tree = make_tree(tab_scenario)
    costs_tree = make_tree(tab_costs)
    impact_note = ttk.Label(
        tab_impacts,
        text=(
            "Risk-conflict hours = baseline working hours where the activity needs more resources "
            "than are available after the selected risk events.\n"
            "Criticality score = downstream chain duration used by the model to prioritize activities; "
            "higher values are closer to the project critical path."
        ),
        justify="left",
        foreground="#333333"
    )
    impact_note.pack(fill="x", pady=(0, 6))
    impacts_tree = make_tree(tab_impacts)

    def make_section(parent, title, note_text):
        section = ttk.LabelFrame(parent, text=title, padding=8)
        section.pack(fill="both", expand=True, pady=(0, 10))
        ttk.Label(section, text=note_text, justify="left", foreground="#333333").pack(fill="x", pady=(0, 6))
        return make_tree(section)

    direct_tree = make_section(
        tab_recovery_details,
        "Direct recovery",
        "Direct recovery details show compressed activities, multipliers, recovered hours, activity-level extra cost %, and project-level extra cost %."
    )
    chain_tree = make_section(
        tab_recovery_details,
        "Cost-optimized chain recovery",
        "Chain recovery details show the optimized recovery plan, including activity-level extra cost % and project-level extra cost %."
    )

    extra_direct_tree = make_section(
        tab_extra_resources,
        "Extra resources - direct recovery",
        "Additional workers/machines required by the direct recovery plan."
    )
    extra_chain_tree = make_section(
        tab_extra_resources,
        "Extra resources - cost-optimized chain recovery",
        "Additional workers/machines required by the optimized chain recovery plan."
    )

    candidates_note = ttk.Label(
        tab_chain_candidates,
        text=(
            "Chain candidates lists all possible recovery actions considered by the optimization, sorted by cost-effectiveness."
        ),
        justify="left",
        foreground="#333333"
    )
    candidates_note.pack(fill="x", pady=(0, 6))
    chain_candidates_tree = make_tree(tab_chain_candidates)

    technical_log_text = tk.Text(tab_log, wrap="none", state="disabled")
    technical_log_text.pack(fill="both", expand=True)

    def choose_existing_columns(df, columns):
        if df is None or getattr(df, "empty", False):
            return df
        selected = [c for c in columns if c in df.columns]
        return df[selected].copy() if selected else df.copy()

    def make_extra_resources_dataframe(df):
        if df is None or getattr(df, "empty", False):
            return pd.DataFrame()
        id_cols = [c for c in [
            "Task ID", "task_id", "Activity", "Activity name", "full_name",
            "Recovery impact type", "Tipo impatto recovery",
            "Baseline multiplier", "Recovery multiplier"
        ] if c in df.columns]
        extra_cols = [c for c in df.columns if str(c).startswith("Extra ")]
        cols = id_cols + extra_cols
        return df[cols].copy() if cols else pd.DataFrame()

    def update_summary_from_results(result_path, lang):
        if lang == "EN":
            comparison = read_sheet_any(result_path, ["Scenario_comparison", "Confronto_scenari"])
            costs = read_sheet_any(result_path, ["Recovery_cost", "Costo_recovery"])
            impacts = read_sheet_any(result_path, ["Impacted_activities", "Attivita_colpite", "Task_colpite"])
            risklog = read_sheet_any(result_path, ["Applied_risk_log", "Log_imprevisti"])
        else:
            comparison = read_sheet_any(result_path, ["Confronto_scenari", "Scenario_comparison"])
            costs = read_sheet_any(result_path, ["Costo_recovery", "Recovery_cost"])
            impacts = read_sheet_any(result_path, ["Attivita_colpite", "Impacted_activities", "Task_colpite"])
            risklog = read_sheet_any(result_path, ["Log_imprevisti", "Applied_risk_log"])

        direct_plan = read_sheet_any(result_path, ["Direct_recovery_plan", "Proposta_recovery_dir"])
        chain_plan = read_sheet_any(result_path, ["Chain_plan", "Proposta_chain"])
        chain_candidates = read_sheet_any(result_path, ["Chain_candidates", "Candidate_chain"])

        def _to_float_safe(v):
            try:
                if pd.isna(v):
                    return None
                if isinstance(v, str):
                    v = v.replace("€", "").replace("%", "").replace(" ", "").strip()
                    # Accept both English 111,425.01 and Italian 111.425,01 formats.
                    if "," in v and "." in v:
                        if v.rfind(",") > v.rfind("."):
                            v = v.replace(".", "").replace(",", ".")
                        else:
                            v = v.replace(",", "")
                    elif "," in v:
                        v = v.replace(",", ".")
                return float(v)
            except Exception:
                return None

        def _first_existing_col(df, candidates):
            if df is None or getattr(df, "empty", False):
                return None
            for c in candidates:
                if c in df.columns:
                    return c
            return None

        def _cost_row_for(costs_df, pattern):
            if costs_df is None or getattr(costs_df, "empty", False):
                return None
            scen_col = _first_existing_col(costs_df, ["Scenario"])
            if scen_col is None:
                return None
            m = costs_df[costs_df[scen_col].astype(str).str.contains(pattern, case=False, na=False)]
            return m.iloc[0] if not m.empty else None

        def add_project_cost_context(plan_df, costs_df, pattern):
            """Add project-level cost context to every recovery-detail row.

            The plan rows already contain activity cost, total extra cost and % extra on activity cost.
            This adds total project cost and % extra on total project cost so the two percentages are explicit.
            """
            if plan_df is None or getattr(plan_df, "empty", False):
                return pd.DataFrame()
            out = plan_df.copy()
            row = _cost_row_for(costs_df, pattern)
            if row is None:
                return out
            total_project_col = _first_existing_col(costs_df, ["Baseline total project cost", "Costo totale baseline progetto"])
            total_extra_col = _first_existing_col(out, ["Total extra cost", "Costo extra totale"])
            total_project = _to_float_safe(row[total_project_col]) if total_project_col else None
            if total_project is not None:
                out["Total project cost"] = total_project
                if total_extra_col:
                    out["% extra on total project cost"] = out[total_extra_col].apply(
                        lambda x: (_to_float_safe(x) / total_project) if total_project else None
                    )
            return out

        # Compact but complete views, matching the original Colab report sections.
        direct_cols = [
            "task_id", "Task ID", "full_name", "Activity", "Activity name",
            "Baseline multiplier", "Recovery multiplier",
            "Durata baseline", "Baseline duration (h)", "Durata recovery", "Recovery duration (h)",
            "Ore recuperate", "Recovered hours (h)", "Costo attività", "Activity cost",
            "Costo totale baseline progetto", "Total project cost",
            "Costo extra totale", "Total extra cost", "% costo extra totale", "% total extra cost", "% extra on activity cost",
            "% extra su totale progetto", "% extra on total project cost",
            "Costo per ora recuperata", "Cost per recovered hour"
        ]
        chain_cols = [
            "task_id", "Task ID", "full_name", "Activity", "Activity name",
            "Tipo impatto recovery", "Recovery impact type", "Regola macchinari", "Machinery rule",
            "Baseline multiplier", "Recovery multiplier",
            "Durata baseline", "Baseline duration (h)", "Durata recovery", "Recovery duration (h)",
            "Ore recuperate", "Recovered hours (h)", "Costo attività", "Activity cost",
            "Costo totale baseline progetto", "Total project cost",
            "Costo extra totale", "Total extra cost", "% costo extra totale", "% total extra cost", "% extra on activity cost",
            "% extra su totale progetto", "% extra on total project cost",
            "Costo per ora recuperata", "Cost per recovered hour"
        ]
        candidate_cols = [
            "task_id", "Task ID", "full_name", "Activity", "Activity name",
            "Tipo impatto recovery", "Recovery impact type", "Regola macchinari", "Machinery rule",
            "Baseline multiplier", "Recovery multiplier", "Ore recuperate", "Recovered hours (h)",
            "Costo extra totale", "Total extra cost", "Costo per ora recuperata", "Cost per recovered hour",
            "Criticality", "Criticality score"
        ]

        set_tree_dataframe(scenario_tree, comparison, lang=lang)
        set_tree_dataframe(costs_tree, costs, lang=lang)
        set_tree_dataframe(impacts_tree, impacts if not impacts.empty else risklog, lang=lang)
        direct_plan_context = add_project_cost_context(direct_plan, costs, "direct|diretta")
        chain_plan_context = add_project_cost_context(chain_plan, costs, "chain|catena")

        set_tree_dataframe(direct_tree, choose_existing_columns(direct_plan_context, direct_cols), lang=lang)
        set_tree_dataframe(extra_direct_tree, make_extra_resources_dataframe(direct_plan), lang=lang)
        set_tree_dataframe(chain_tree, choose_existing_columns(chain_plan_context, chain_cols), lang=lang)
        set_tree_dataframe(extra_chain_tree, make_extra_resources_dataframe(chain_plan), lang=lang)
        set_tree_dataframe(chain_candidates_tree, choose_existing_columns(chain_candidates, candidate_cols), lang=lang)

        try:
            if not comparison.empty:
                scen_col = "Scenario"
                finish_col = "Project finish (h)" if "Project finish (h)" in comparison.columns else "Fine progetto (ora assoluta)"
                delay_h_col = "Delay vs baseline (h)" if "Delay vs baseline (h)" in comparison.columns else "Ritardo vs baseline (h)"
                delay_d_col = "Delay vs baseline (work days)" if "Delay vs baseline (work days)" in comparison.columns else "Ritardo vs baseline (gg lav)"

                def fmt_num(x):
                    try:
                        xf = float(x)
                        return str(int(xf)) if xf.is_integer() else f"{xf:g}"
                    except Exception:
                        return str(x)

                def fmt_h(x):
                    return f"{fmt_num(x)} h"

                def find_row_contains(txt):
                    m = comparison[comparison[scen_col].astype(str).str.contains(txt, case=False, na=False)]
                    return m.iloc[0] if not m.empty else None
                base = find_row_contains("Baseline")
                risks = find_row_contains("no recovery|senza recovery")
                direct = find_row_contains("direct|diretta")
                chain = find_row_contains("chain|catena")
                if base is not None: summary_vars["baseline"].set(fmt_h(base[finish_col]))
                if risks is not None:
                    summary_vars["risk_finish"].set(fmt_h(risks[finish_col]))
                    summary_vars["risk_delay"].set(f"{fmt_h(risks[delay_h_col])} ({fmt_num(risks[delay_d_col])} work days)")
                if direct is not None: summary_vars["direct_finish"].set(fmt_h(direct[finish_col]))
                if chain is not None: summary_vars["chain_finish"].set(fmt_h(chain[finish_col]))
        except Exception:
            pass

        try:
            direct_cost = money_value(costs, "direct|diretta", ["Total recovery extra cost", "Costo extra totale recovery"])
            chain_cost = money_value(costs, "chain|catena", ["Total recovery extra cost", "Costo extra totale recovery"])
            if direct_cost is not None: summary_vars["direct_cost"].set(str(format_euro_value(direct_cost)))
            if chain_cost is not None: summary_vars["chain_cost"].set(str(format_euro_value(chain_cost)))
        except Exception:
            pass

        summary_vars["output"].set(str(result_path))

    def run_worker():
        old_stdout, old_stderr = sys.stdout, sys.stderr
        result_path = None
        try:
            lang = "EN" if output_language_var.get() == "English" else "IT"
            global IMPREVISTI_INPUT
            IMPREVISTI_INPUT = list(imprevisti_rows)
            input_path = str(resolve_input_excel_path(input_var.get().strip()))
            output_path = output_var.get().strip()

            app.after(0, lambda: set_text(log_clean, "Simulation running...\n"))
            app.after(0, lambda: set_text(technical_log_text, ""))

            buf = io.StringIO()
            sys.stdout = buf
            sys.stderr = buf
            result_path = run_simulation(input_xlsx=input_path, output_xlsx=output_path, no_prompt=True, terminal_prompt=False)
            sys.stdout, sys.stderr = old_stdout, old_stderr

            raw_log = buf.getvalue()
            display_log = translate_runtime_log_to_english(raw_log) if lang == "EN" else raw_log
            if lang == "EN":
                postprocess_output_workbook_language(result_path, "EN")

            output_created_path.set(str(result_path))

            clean_lines = [
                "Simulation completed successfully.",
                f"Input: {input_path}",
                f"Output: {result_path}",
                f"Risk events: {len(IMPREVISTI_INPUT)}",
                f"Output workbook language: {'English' if lang == 'EN' else 'Italian'}",
                "",
                "Use the tabs above to inspect scenario comparison, recovery costs, recovery details, extra resources and impacted activities.",
            ] if lang == "EN" else [
                "Simulazione completata correttamente.",
                f"Input: {input_path}",
                f"Output: {result_path}",
                f"Imprevisti: {len(IMPREVISTI_INPUT)}",
                f"Lingua output: {'Inglese' if lang == 'EN' else 'Italiano'}",
                "",
                "Usa le schede sopra per vedere confronto scenari, costi, dettagli recovery, risorse extra e attività colpite.",
            ]

            app.after(0, lambda: set_text(log_clean, "\n".join(clean_lines)))
            app.after(0, lambda: set_text(technical_log_text, display_log))
            app.after(0, lambda: update_summary_from_results(result_path, lang))
            app.after(0, lambda: status_var.set("Simulation completed."))
            app.after(0, lambda: open_output_btn.configure(state="normal"))
            app.after(0, lambda: nb.select(tab_summary))

        except Exception:
            sys.stdout, sys.stderr = old_stdout, old_stderr
            err = traceback.format_exc()
            app.after(0, lambda: set_text(log_clean, "Simulation failed. See the Technical log tab for details."))
            app.after(0, lambda: set_text(technical_log_text, err))
            app.after(0, lambda: status_var.set("Simulation failed."))
            app.after(0, lambda: messagebox.showerror("Simulation error", "The simulation failed. Check the Technical log tab for details."))
        finally:
            sys.stdout, sys.stderr = old_stdout, old_stderr
            app.after(0, lambda: progress.stop())
            app.after(0, lambda: run_btn.configure(state="normal"))
            app.after(0, lambda: is_running.set(False))

    def run_from_gui():
        if is_running.get():
            return
        input_path = resolve_input_excel_path(input_var.get().strip())
        if not input_path.exists():
            messagebox.showerror("Input file not found", f"Input Excel file not found:\n{input_path}\n\nUse Browse to select it.")
            return
        input_var.set(str(input_path))
        if not imprevisti_rows:
            messagebox.showerror("Missing risk events", "Add at least one risk event before running the simulation.")
            return
        if not output_var.get().strip():
            messagebox.showerror("Missing output file", "Choose where to save the output Excel file.")
            return
        is_running.set(True)
        run_btn.configure(state="disabled")
        open_output_btn.configure(state="disabled")
        progress.start(10)
        status_var.set("Simulation running...")
        threading.Thread(target=run_worker, daemon=True).start()

    lang_combo.bind("<<ComboboxSelected>>", lambda _e: refresh_risk_table())
    app.after(250, lambda: load_calendar_values(show_errors=False))
    app.mainloop()


def main():
    parser = argparse.ArgumentParser(description="Construction risk simulation tool.")
    parser.add_argument("--input", "-i", default=INPUT_XLSX, help="Input Excel file with PL and Orario sheets.")
    parser.add_argument("--output", "-o", default=OUTPUT_XLSX, help="Output Excel file to create.")
    parser.add_argument("--terminal", action="store_true", help="Use classic terminal mode instead of the GUI.")
    parser.add_argument("--no-prompt", action="store_true", help="Do not ask for risk events; use IMPREVISTI_INPUT only.")
    parser.add_argument("--list-risks", action="store_true", help="Show available risk events and exit.")
    args = parser.parse_args()

    if args.list_risks:
        print("Available risk events:")
        for name, data in RISK_LIBRARY.items():
            print(f"- {name}: {data}")
        return

    if args.terminal or args.no_prompt:
        run_simulation(args.input, args.output, no_prompt=args.no_prompt, terminal_prompt=args.terminal)
    else:
        launch_full_gui(default_input=args.input, default_output=args.output)


if __name__ == "__main__":
    main()
